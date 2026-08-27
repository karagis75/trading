import argparse
import logging
import math
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import requests

try:
    import yfinance as yf
except ImportError:
    yf = None

from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from curl_cffi import requests as curl_requests
except ImportError:
    curl_requests = None


NSE_HOME_URL = "https://www.nseindia.com"
INDICES_URL = "https://www.nseindia.com/api/allIndices"
OPTION_CHAIN_CONTRACT_INFO_URL = "https://www.nseindia.com/api/option-chain-contract-info"
OPTION_CHAIN_URL = "https://www.nseindia.com/api/option-chain-v3"

DEFAULT_SYMBOLS = (
    "BIOCON",
    "ADANIPORTS",
    "INFY",
    "HDFCBANK",
    "ICICIBANK",
    "SBIN",
    "AUBANK",
    "BHARTIARTL",
    "BHEL",
    "HINDPETRO",
    "IOC"
)

INDEX_SYMBOLS = {"NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "NIFTYNXT50"}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
    "DNT": "1",
    "Referer": NSE_HOME_URL,
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Upgrade-Insecure-Requests": "1",
}

# Fast O(1) strategy directional mapping for confluence verification
STRATEGY_DIRECTION_MAP = {
    "Bull Call Spread": "BULLISH",
    "Bull Put Credit Spread": "BULLISH",
    "Bear Put Spread": "BEARISH",
    "Bear Call Credit Spread": "BEARISH",
    "Iron Condor": "NEUTRAL",
    "Short Iron Butterfly": "NEUTRAL",
    "Long Strangle": "SIDEWAYS",
    "Short Straddle": "NEUTRAL",
}

LOGGER = logging.getLogger("combined_option_scanner")


@dataclass(frozen=True)
class ScannerConfig:
    min_bid: float = 0.05
    min_open_interest: int = 100
    max_legs_apart: int = 8
    atm_window_pct: float = 0.12
    neutral_pcr_low: float = 0.7
    neutral_pcr_high: float = 1.3
    fallback_vix: float = 15.0
    min_credit_pct_of_width: float = 0.12
    min_short_distance_pct: float = 0.01
    max_bid_ask_spread_pct: float = 0.25
    min_volume: int = 10
    max_margin_per_trade: float | None = None
    lot_size: int = 1
    enforce_validations: bool = False
    lookback_period: str = "1y"


@dataclass(frozen=True)
class MarketContext:
    symbol: str
    records: list[dict[str, Any]]
    underlying_price: float
    pcr: float
    max_open_interest: int
    expiry: str
    trend: str = "unknown"
    event_risk: str = "unknown"
    atm_iv: float = 0.0  # decimal (e.g. 0.18 for 18%), derived from symbol chain


def effective_iv(context: "MarketContext", india_vix: float) -> float:
    if context.atm_iv and context.atm_iv > 0:
        return context.atm_iv
    return max(0.01, india_vix / 100.0)


# ---------------------------------------------------------------------------
# Technical Trend Scanners Logic
# ---------------------------------------------------------------------------

def calculate_technical_indicators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Flatten MultiIndex columns if present (from yfinance)
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [col[0] for col in df.columns]

    # Clean missing OHLC data & derive Close if NaN
    df[["Open", "High", "Low", "Close"]] = df[["Open", "High", "Low", "Close"]].ffill().bfill()
    if df["Close"].isna().any():
        df["Close"] = df["Close"].fillna((df["High"] + df["Low"]) / 2).fillna(df["Open"])

    # Ensure required columns are float64
    for col in ["Open", "High", "Low", "Close"]:
        if col in df.columns:
            df[col] = df[col].astype(float)

    # EMAs
    df["EMA9"] = df["Close"].ewm(span=9, adjust=False).mean()
    df["EMA18"] = df["Close"].ewm(span=18, adjust=False).mean()
    df["EMA50"] = df["Close"].ewm(span=50, adjust=False).mean()
    df["EMA200"] = df["Close"].ewm(span=200, adjust=False).mean()

    # Robust CCI (14 & 20) with min_periods and zero-MAD replacement
    tp = (df["High"] + df["Low"] + df["Close"]) / 3.0
    
    sma_tp14 = tp.rolling(window=14, min_periods=1).mean()
    mad14 = tp.rolling(window=14, min_periods=1).apply(lambda x: np.abs(x - x.mean()).mean(), raw=True).replace(0, np.nan)
    df["CCI14"] = ((tp - sma_tp14) / (0.015 * mad14)).fillna(0)

    sma_tp20 = tp.rolling(window=20, min_periods=1).mean()
    mad20 = tp.rolling(window=20, min_periods=1).apply(lambda x: np.abs(x - x.mean()).mean(), raw=True).replace(0, np.nan)
    df["CCI20"] = ((tp - sma_tp20) / (0.015 * mad20)).fillna(0)

    # Robust ADX (14)
    high_diff = df["High"].diff()
    low_diff = -df["Low"].diff()
    pos_dm = np.where((high_diff > low_diff) & (high_diff > 0), high_diff, 0.0)
    neg_dm = np.where((low_diff > high_diff) & (low_diff > 0), low_diff, 0.0)

    tr1 = df["High"] - df["Low"]
    tr2 = (df["High"] - df["Close"].shift(1)).abs()
    tr3 = (df["Low"] - df["Close"].shift(1)).abs()
    tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)

    atr = tr.rolling(window=14, min_periods=1).mean().replace(0, np.nan)
    pos_di = 100 * (pd.Series(pos_dm, index=df.index).rolling(window=14, min_periods=1).mean() / atr)
    neg_di = 100 * (pd.Series(neg_dm, index=df.index).rolling(window=14, min_periods=1).mean() / atr)
    
    di_sum = (pos_di + neg_di).replace(0, np.nan)
    dx = 100 * ((pos_di - neg_di).abs() / di_sum)
    df["ADX"] = dx.rolling(window=14, min_periods=1).mean().fillna(0)

    # Fibonacci S1
    prev_high = df["High"].shift(1).fillna(df["High"])
    prev_low = df["Low"].shift(1).fillna(df["Low"])
    prev_close = df["Close"].shift(1).fillna(df["Close"])
    pivot = (prev_high + prev_low + prev_close) / 3.0
    df["Fib_S1"] = pivot - 0.382 * (prev_high - prev_low)

    return df


def detect_market_trend(symbol: str, lookback_period: str = "1y") -> str:
    trend, _ = detect_market_trend_detailed(symbol, lookback_period)
    return trend


def detect_market_trend_detailed(
    symbol: str, lookback_period: str = "1y"
) -> tuple[str, dict[str, Any]]:
    diag: dict[str, Any] = {"Symbol": symbol, "Data Rows": 0, "Fetch Error": ""}

    if yf is None:
        LOGGER.warning("yfinance package not installed. Skipping dynamic trend detection.")
        diag["Fetch Error"] = "yfinance not installed"
        return "unknown", diag

    formatted_ticker = symbol if ("." in symbol or symbol.startswith("^")) else f"{symbol}.NS"
    diag["Ticker Used"] = formatted_ticker
    try:
        ticker = yf.Ticker(formatted_ticker)
        df = ticker.history(period=lookback_period, interval="1d")
        
        if df.empty:
            diag["Fetch Error"] = "No data returned from yfinance"
            return "unknown", diag

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [col[0] for col in df.columns]

        diag["Data Rows"] = len(df)

        if len(df) < 50:
            diag["Fetch Error"] = f"Insufficient history ({len(df)} rows, need >=50)"
            return "unknown", diag

        df = calculate_technical_indicators(df)
        curr = df.iloc[-1]
        prev = df.iloc[-2]

        diag.update({
            "Close": round(float(curr["Close"]), 2) if pd.notna(curr["Close"]) else None,
            "EMA9": round(float(curr["EMA9"]), 2) if pd.notna(curr["EMA9"]) else None,
            "EMA18": round(float(curr["EMA18"]), 2) if pd.notna(curr["EMA18"]) else None,
            "EMA50": round(float(curr["EMA50"]), 2) if pd.notna(curr["EMA50"]) else None,
            "EMA200": round(float(curr["EMA200"]), 2) if len(df) >= 200 and pd.notna(curr["EMA200"]) else None,
            "CCI14": round(float(curr["CCI14"]), 2) if pd.notna(curr["CCI14"]) else None,
            "CCI20": round(float(curr["CCI20"]), 2) if pd.notna(curr["CCI20"]) else None,
            "ADX": round(float(curr["ADX"]), 2) if pd.notna(curr["ADX"]) else None,
            "Fib_S1": round(float(curr["Fib_S1"]), 2) if pd.notna(curr["Fib_S1"]) else None,
        })

        # 1. Bullish Setup
        bullish_ready = len(df) >= 200
        diag["Bullish: enough history (>=200d)"] = bullish_ready
        if bullish_ready and pd.notna(curr["EMA200"]):
            price_above_emas = (
                curr["Close"] > curr["EMA9"]
                and curr["EMA9"] > curr["EMA18"]
                and curr["EMA18"] > curr["EMA50"]
                and curr["EMA50"] > curr["EMA200"]
            )
            ema200_sloping_up = curr["EMA200"] >= df.iloc[-5]["EMA200"]
            cci_trigger = (prev["CCI20"] <= 100 and curr["CCI20"] > 100) or (curr["CCI20"] > 100)
            body_size = abs(curr["Close"] - curr["Open"])
            candle_range = curr["High"] - curr["Low"]
            is_bullish_candle = (curr["Close"] > curr["Open"]) and (
                body_size / candle_range >= 0.5 if candle_range > 0 else True
            )
            riding_ema9 = curr["Low"] >= curr["EMA9"] or curr["Close"] > curr["EMA9"]
            adx_pass_bull = pd.notna(curr["ADX"]) and curr["ADX"] >= 20.0

            diag["Bullish: EMA stack (P>9>18>50>200)"] = bool(price_above_emas)
            diag["Bullish: EMA200 sloping up"] = bool(ema200_sloping_up)
            diag["Bullish: CCI20 > 100 trigger"] = bool(cci_trigger)
            diag["Bullish: bullish candle body>=50%"] = bool(is_bullish_candle)
            diag["Bullish: riding EMA9"] = bool(riding_ema9)
            diag["Bullish: ADX >= 20"] = bool(adx_pass_bull)

            if (
                price_above_emas and ema200_sloping_up and cci_trigger
                and is_bullish_candle and riding_ema9 and adx_pass_bull
            ):
                diag["Result"] = "bullish"
                return "bullish", diag

        # 2. Bearish Setup
        ema_alignment_bearish = (
            curr["EMA50"] > curr["EMA18"]
            and curr["EMA18"] > curr["EMA9"]
            and curr["Close"] < curr["EMA9"]
        )
        adx_pass_bear = pd.notna(curr["ADX"]) and curr["ADX"] >= 20.0
        cci_pass_bear = pd.notna(curr["CCI14"]) and curr["CCI14"] <= -100.0
        fib_pass = pd.notna(curr["Fib_S1"]) and curr["Close"] < curr["Fib_S1"]

        diag["Bearish: ADX >= 20"] = bool(adx_pass_bear)
        diag["Bearish: CCI14 <= -100"] = bool(cci_pass_bear)
        diag["Bearish: EMA stack (50>18>9>Price)"] = bool(ema_alignment_bearish)
        diag["Bearish: Close < Fib S1"] = bool(fib_pass)

        if adx_pass_bear and cci_pass_bear and ema_alignment_bearish and fib_pass:
            diag["Result"] = "bearish"
            return "bearish", diag

        # 3. Rangebound / Strangle Compression Setup
        recent_df = df.iloc[-3:]
        adx_below_threshold = (recent_df["ADX"] < 15.0).any()
        cci_tight = pd.notna(curr["CCI14"]) and (-50.0 <= curr["CCI14"] <= 50.0)

        ema_min = min(curr["EMA9"], curr["EMA18"], curr["EMA50"])
        ema_max = max(curr["EMA9"], curr["EMA18"], curr["EMA50"])
        ema_spread_pct = (ema_max - ema_min) / curr["Close"]
        ema_braid_pass = ema_spread_pct <= 0.012

        box_df = df.iloc[-6:]
        box_high = box_df["High"].max()
        box_low = box_df["Low"].min()
        box_range_pct = (box_high - box_low) / curr["Close"]
        box_pass = box_range_pct <= 0.035

        diag["Sideways: ADX < 15 (last 3d)"] = bool(adx_below_threshold)
        diag["Sideways: CCI14 in [-50,50]"] = bool(cci_tight)
        diag["Sideways: EMA braid <=1.2%"] = bool(ema_braid_pass)
        diag["Sideways: EMA braid spread %"] = round(float(ema_spread_pct) * 100, 2)
        diag["Sideways: 6d box range <=3.5%"] = bool(box_pass)
        diag["Sideways: 6d box range %"] = round(float(box_range_pct) * 100, 2)

        if adx_below_threshold and cci_tight and ema_braid_pass and box_pass:
            diag["Result"] = "sideways"
            return "sideways", diag

        diag["Result"] = "neutral"
        return "neutral", diag

    except Exception as exc:
        LOGGER.warning("Failed to calculate trend for %s: %s", symbol, exc)
        diag["Fetch Error"] = str(exc)
        return "unknown", diag


# ---------------------------------------------------------------------------
# NSE API & Cookie Handling
# ---------------------------------------------------------------------------

def load_cookie_header(cookie_header: str | None, cookie_file: str | None) -> str | None:
    if cookie_header:
        return cookie_header.strip()
    if cookie_file:
        return Path(cookie_file).read_text(encoding="utf-8").strip()
    return None


def create_nse_session(
    cookie_header: str | None = None,
    use_browser_impersonation: bool = False,
) -> Any:
    if use_browser_impersonation:
        if curl_requests is None:
            raise RuntimeError("curl_cffi is not installed. Run: pip install curl_cffi")
        session = curl_requests.Session(impersonate="chrome120")
    else:
        session = requests.Session()

    session.headers.update(HEADERS)
    if cookie_header:
        session.headers.update({"Cookie": cookie_header})

    if not use_browser_impersonation:
        retry = Retry(
            total=3, connect=3, read=3, backoff_factor=0.5,
            status_forcelist=(429, 500, 502, 503, 504), allowed_methods=("GET",)
        )
        session.mount("https://", HTTPAdapter(max_retries=retry))

    response = session.get(NSE_HOME_URL, timeout=10)
    response.raise_for_status()
    session.get(f"{NSE_HOME_URL}/option-chain", timeout=10)
    return session


def get_json(session: Any, url: str, **params: Any) -> dict[str, Any] | None:
    api_headers = {
        **HEADERS,
        "Accept": "application/json,text/plain,*/*",
        "Referer": f"{NSE_HOME_URL}/option-chain",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "Upgrade-Insecure-Requests": "0",
    }
    try:
        response = session.get(url, params=params or None, timeout=15, headers=api_headers)
        if response.status_code == 403:
            LOGGER.warning("NSE returned 403 for %s.", url)
            return None
        response.raise_for_status()
        return response.json()
    except Exception as exc:
        LOGGER.warning("Request failed for %s: %s", url, exc)
    return None


def fetch_india_vix(session: Any, fallback: float) -> float:
    payload = get_json(session, INDICES_URL)
    for index in (payload or {}).get("data", []):
        if index.get("index") == "INDIA VIX":
            try:
                return float(index["last"])
            except (KeyError, TypeError, ValueError):
                break
    return fallback


def fetch_contract_info(session: Any, symbol: str) -> dict[str, Any] | None:
    return get_json(session, OPTION_CHAIN_CONTRACT_INFO_URL, symbol=symbol.upper())


def detect_chain_type(symbol: str, chain_type: str) -> str:
    if chain_type != "auto":
        return "Indices" if chain_type == "indices" else "Equity"
    return "Indices" if symbol.upper() in INDEX_SYMBOLS else "Equity"


def fetch_option_chain(
    session: Any, symbol: str, chain_type: str, expiry: str | None = None
) -> dict[str, Any] | None:
    symbol = symbol.upper()
    contract_info = fetch_contract_info(session, symbol)
    expiry_dates = (contract_info or {}).get("expiryDates") or []
    if expiry and expiry_dates and expiry not in expiry_dates:
        LOGGER.warning("Expiry %s is not available for %s.", expiry, symbol)
        return None
    selected_expiry = expiry or (expiry_dates[0] if expiry_dates else None)
    if not selected_expiry:
        return None

    payload = get_json(
        session,
        OPTION_CHAIN_URL,
        type=detect_chain_type(symbol, chain_type),
        symbol=symbol,
        expiry=selected_expiry,
    )
    if payload:
        payload["_selected_expiry"] = selected_expiry
    return payload


def latest_expiry_records(data: dict[str, Any]) -> list[dict[str, Any]]:
    records = data.get("records", {}).get("data") or []
    selected_expiry = data.get("_selected_expiry")
    if not selected_expiry:
        return records

    matching_records = [
        row for row in records
        if row.get("expiryDate") == selected_expiry
        or selected_expiry in (row.get("expiryDates") or [])
    ]
    # Some NSE responses already contain only the requested expiry and omit the
    # expiry field from each row. Preserve those responses rather than dropping
    # the entire chain.
    return matching_records or records


def load_symbol_map(path: str | None, value_column: str) -> dict[str, str]:
    if not path:
        return {}
    df = pd.read_csv(path)
    normalized_columns = {"".join(str(col).lower().split()): col for col in df.columns}
    symbol_column = normalized_columns.get("symbol")
    value_column_name = normalized_columns.get("".join(value_column.lower().split()))
    if not symbol_column or not value_column_name:
        raise ValueError(f"{path} must contain Symbol and {value_column} columns")
    return {
        str(row[symbol_column]).upper(): str(row[value_column_name]).strip()
        for _, row in df.iterrows()
        if pd.notna(row[symbol_column]) and pd.notna(row[value_column_name])
    }


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def bid_price(option: dict[str, Any]) -> float:
    return _safe_float(option.get("bidprice") or option.get("buyPrice1"))


def ask_price(option: dict[str, Any]) -> float:
    return _safe_float(option.get("askPrice") or option.get("sellPrice1"))


def traded_volume(option: dict[str, Any]) -> int:
    return _safe_int(option.get("totalTradedVolume") or option.get("volume"))


def bid_ask_spread_pct(option: dict[str, Any]) -> float:
    ask = ask_price(option)
    return max(0.0, ask - bid_price(option)) / ask if ask > 0 else 1.0


def option_is_tradeable(option: dict[str, Any], config: ScannerConfig) -> bool:
    return (
        bid_price(option) >= config.min_bid
        and ask_price(option) > 0
        and _safe_int(option.get("openInterest")) >= config.min_open_interest
        and traded_volume(option) >= config.min_volume
        and bid_ask_spread_pct(option) <= config.max_bid_ask_spread_pct
    )


def build_market_context(
    data: dict[str, Any],
    symbol: str,
    expiry: str | None,
    trend_map: dict[str, str],
    event_map: dict[str, str],
    config: ScannerConfig,
) -> MarketContext | None:
    records = latest_expiry_records(data)
    if not records:
        return None

    underlying_price = _safe_float(data["records"].get("underlyingValue"))
    if underlying_price <= 0:
        for row in records:
            underlying_price = _safe_float(
                row.get("CE", {}).get("underlyingValue")
                or row.get("PE", {}).get("underlyingValue")
            )
            if underlying_price > 0:
                break
    if underlying_price <= 0:
        return None

    total_ce_oi = sum(_safe_int(row.get("CE", {}).get("openInterest")) for row in records)
    total_pe_oi = sum(_safe_int(row.get("PE", {}).get("openInterest")) for row in records)
    pcr = total_pe_oi / total_ce_oi if total_ce_oi > 0 else 1.0
    max_open_interest = max(
        [_safe_int(row.get("CE", {}).get("openInterest")) for row in records]
        + [_safe_int(row.get("PE", {}).get("openInterest")) for row in records]
        + [1]
    )

    symbol_key = symbol.upper()
    trend = trend_map.get(symbol_key, "").lower()

    if not trend or trend == "unknown":
        trend = detect_market_trend(symbol_key, lookback_period=config.lookback_period)

    atm_iv = 0.0
    closest_row = min(
        records,
        key=lambda row: abs(
            _safe_float(
                (row.get("CE") or {}).get("strikePrice")
                or (row.get("PE") or {}).get("strikePrice")
            )
            - underlying_price
        ),
    )
    ce_iv = _safe_float((closest_row.get("CE") or {}).get("impliedVolatility"))
    pe_iv = _safe_float((closest_row.get("PE") or {}).get("impliedVolatility"))
    ivs = [v for v in (ce_iv, pe_iv) if v > 0]
    if ivs:
        atm_iv = (sum(ivs) / len(ivs)) / 100.0

    return MarketContext(
        symbol=symbol_key,
        records=records,
        underlying_price=underlying_price,
        pcr=pcr,
        max_open_interest=max_open_interest,
        expiry=data.get("_selected_expiry") or expiry or "",
        trend=trend,
        event_risk=event_map.get(symbol_key, "unknown").lower(),
        atm_iv=atm_iv,
    )


def trend_allows_strategy(strategy: str, trend: str) -> bool:
    if trend in ("unknown", ""):
        return True

    required_direction = STRATEGY_DIRECTION_MAP.get(strategy)
    if not required_direction:
        return True

    trend_upper = trend.upper()
    if required_direction == "BULLISH":
        return trend_upper == "BULLISH"
    elif required_direction == "BEARISH":
        return trend_upper == "BEARISH"
    elif required_direction == "NEUTRAL":
        return trend_upper in ("NEUTRAL", "SIDEWAYS")
    elif required_direction == "SIDEWAYS":
        return trend_upper in ("SIDEWAYS", "NEUTRAL")

    return False


def pcr_bias(context: MarketContext, config: ScannerConfig) -> str:
    if context.pcr < config.neutral_pcr_low:
        return "bullish"
    if context.pcr > config.neutral_pcr_high:
        return "bearish"
    return "neutral"


def pcr_allows_strategy(strategy: str, bias: str) -> bool:
    required_direction = STRATEGY_DIRECTION_MAP.get(strategy)
    if not required_direction:
        return True

    bias_upper = bias.upper()
    if required_direction == "BULLISH":
        return bias_upper == "BULLISH"
    elif required_direction == "BEARISH":
        return bias_upper == "BEARISH"
    elif required_direction in ("NEUTRAL", "SIDEWAYS"):
        return bias_upper == "NEUTRAL"

    return False


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def add_validation_fields(
    opportunity: dict[str, Any] | None, context: MarketContext, config: ScannerConfig
) -> dict[str, Any] | None:
    if opportunity is None:
        return None
    strategy = str(opportunity["Strategy"])
    bias = pcr_bias(context, config)

    bias_pass = pcr_allows_strategy(strategy, bias)
    if config.enforce_validations and not bias_pass:
        return None

    raw_max_loss = opportunity.get("Max Loss")
    if isinstance(raw_max_loss, str) and raw_max_loss.strip().lower() == "unlimited":
        margin_per_unit = round(context.underlying_price * 0.20, 2)
    else:
        margin_per_unit = (
            _safe_float(raw_max_loss)
            or _safe_float(opportunity.get("Net Debit"))
            or _safe_float(opportunity.get("Credit"))
            or 0.0
        )
    estimated_margin = round(margin_per_unit * config.lot_size, 2)
    spread_width = opportunity.get("Spread Width") or 0
    liquidity_pass = (
        float(opportunity["Avg OI"]) >= config.min_open_interest
        and (
            float(spread_width) <= 0
            or float(opportunity["Bid-Ask Spread"]) <= float(spread_width) * config.max_bid_ask_spread_pct
        )
    )
    trend_pass = trend_allows_strategy(strategy, context.trend)
    event_pass = context.event_risk not in ("yes", "true", "high", "blocked", "avoid")
    margin_pass = (
        config.max_margin_per_trade is None or estimated_margin <= config.max_margin_per_trade
    )

    validation_pass = bias_pass and trend_pass and liquidity_pass and event_pass and margin_pass

    strategy_direction = STRATEGY_DIRECTION_MAP.get(strategy)
    confluence_bonus = 0.0
    if strategy_direction:
        if bias.upper() == strategy_direction or (
            strategy_direction in ("NEUTRAL", "SIDEWAYS") and bias.upper() == "NEUTRAL"
        ):
            confluence_bonus += 4.0
        trend_upper = context.trend.upper()
        trend_matches = (
            trend_upper == strategy_direction
            or (strategy_direction in ("NEUTRAL", "SIDEWAYS") and trend_upper in ("NEUTRAL", "SIDEWAYS"))
        )
        if trend_matches:
            confluence_bonus += 6.0

    opportunity["Score"] = round(min(100.0, float(opportunity["Score"]) + confluence_bonus), 2)

    opportunity.update(
        {
            "Trend": context.trend,
            "Trend Pass": trend_pass,
            "PCR Bias": bias,
            "PCR Bias Pass": bias_pass,
            "Confluence Bonus": confluence_bonus,
            "Event Risk": context.event_risk,
            "Event Risk Pass": event_pass,
            "Liquidity Pass": liquidity_pass,
            "Estimated Margin": estimated_margin,
            "Margin Pass": margin_pass,
            "Validation Pass": validation_pass,
        }
    )

    if config.enforce_validations and not validation_pass:
        return None

    return opportunity


# ---------------------------------------------------------------------------
# Scoring and Opportunity Builders
# ---------------------------------------------------------------------------

def score_spread(
    rr_ratio: float, cost_efficiency: float, avg_open_interest: float,
    max_open_interest: int, total_bid_ask_spread: float, prob_of_profit: float = 0.5
) -> float:
    rr_score = min(30.0, (rr_ratio / 3.0) * 30.0) if rr_ratio > 0 else 0.0
    cost_efficiency_score = max(0.0, (1.0 - cost_efficiency) * 20.0)
    pop_score = min(25.0, max(0.0, prob_of_profit) / 0.6 * 25.0)
    oi_score = (avg_open_interest / max_open_interest * 15.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 10.0 - min(10.0, total_bid_ask_spread * 2.0))
    return round(rr_score + cost_efficiency_score + pop_score + oi_score + tightness_score, 2)


def parse_nse_expiry(expiry: str) -> date | None:
    for fmt in ("%d-%b-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(expiry, fmt).date()
        except ValueError:
            continue
    return None


def days_to_expiry(expiry: str) -> int:
    expiry_date = parse_nse_expiry(expiry)
    return max(1, (expiry_date - date.today()).days) if expiry_date else 1


def expected_move_pct(iv_decimal: float, expiry: str) -> float:
    return max(0.005, iv_decimal * (days_to_expiry(expiry) / 365) ** 0.5)


def _norm_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def probability_otm(
    underlying_price: float, strike: float, iv_decimal: float, dte_days: int, option_type: str
) -> float:
    if underlying_price <= 0 or strike <= 0 or iv_decimal <= 0 or dte_days <= 0:
        return 0.5
    t = dte_days / 365.0
    try:
        d1 = (math.log(underlying_price / strike) + 0.5 * iv_decimal**2 * t) / (iv_decimal * math.sqrt(t))
        d2 = d1 - iv_decimal * math.sqrt(t)
    except (ValueError, ZeroDivisionError):
        return 0.5
    if option_type.upper() == "PUT":
        return _norm_cdf(d2)
    return _norm_cdf(-d2)


def probability_itm(
    underlying_price: float, strike: float, iv_decimal: float, dte_days: int, option_type: str
) -> float:
    return 1.0 - probability_otm(underlying_price, strike, iv_decimal, dte_days, option_type)


def score_credit_spread(
    return_on_risk: float, prob_of_profit: float,
    avg_open_interest: float, max_open_interest: int, total_bid_ask_spread: float,
    credit_pct_of_width: float
) -> float:
    return_score = min(25.0, return_on_risk / 0.35 * 25.0)
    pop_score = min(30.0, max(0.0, prob_of_profit) / 0.85 * 30.0)
    oi_score = (avg_open_interest / max_open_interest * 20.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 15.0 - min(15.0, total_bid_ask_spread * 2.0))
    credit_score = min(10.0, credit_pct_of_width / 0.25 * 10.0)
    return round(return_score + pop_score + oi_score + tightness_score + credit_score, 2)


def build_opportunity(
    symbol: str, strategy: str, pcr: float, india_vix: float, underlying_price: float,
    buy_leg: dict[str, Any], sell_leg: dict[str, Any], max_open_interest: int, expiry: str,
    iv_decimal: float = 0.0,
) -> dict[str, Any] | None:
    buy_price = ask_price(buy_leg)
    sell_price = bid_price(sell_leg)
    net_debit = buy_price - sell_price
    spread_width = abs(float(sell_leg["strikePrice"]) - float(buy_leg["strikePrice"]))

    if net_debit <= 0 or spread_width <= net_debit:
        return None

    max_profit = spread_width - net_debit
    rr_ratio = max_profit / net_debit
    cost_efficiency = net_debit / spread_width
    avg_open_interest = (_safe_int(buy_leg.get("openInterest")) + _safe_int(sell_leg.get("openInterest"))) / 2
    bid_ask_spread = (ask_price(buy_leg) - bid_price(buy_leg) + ask_price(sell_leg) - bid_price(sell_leg))

    dte = days_to_expiry(expiry)
    iv_for_calc = iv_decimal if iv_decimal > 0 else max(0.01, india_vix / 100.0)
    if strategy == "Bull Call Spread":
        breakeven = float(buy_leg["strikePrice"]) + net_debit
        prob_of_profit = probability_itm(underlying_price, breakeven, iv_for_calc, dte, "CALL")
    else:
        breakeven = float(buy_leg["strikePrice"]) - net_debit
        prob_of_profit = probability_itm(underlying_price, breakeven, iv_for_calc, dte, "PUT")

    return {
        "Symbol": symbol,
        "Strategy": strategy,
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Buy Leg (Strike)": buy_leg["strikePrice"],
        "Sell Leg (Strike)": sell_leg["strikePrice"],
        "Net Debit": round(net_debit, 2),
        "Spread Width": round(spread_width, 2),
        "Max Profit": round(max_profit, 2),
        "Max Loss": round(net_debit, 2),
        "R:R Ratio": round(rr_ratio, 2),
        "Breakeven": round(breakeven, 2),
        "Probability of Profit": round(prob_of_profit, 3),
        "Avg OI": round(avg_open_interest),
        "Bid-Ask Spread": round(bid_ask_spread, 2),
        "Score": score_spread(
            rr_ratio, cost_efficiency, avg_open_interest, max_open_interest, bid_ask_spread, prob_of_profit
        ),
    }


def build_credit_spread_opportunity(
    symbol: str, strategy: str, pcr: float, india_vix: float, underlying_price: float,
    short_leg: dict[str, Any], long_leg: dict[str, Any], max_open_interest: int,
    expiry: str, config: ScannerConfig, iv_decimal: float = 0.0,
) -> dict[str, Any] | None:
    short_strike = float(short_leg["strikePrice"])
    long_strike = float(long_leg["strikePrice"])
    spread_width = abs(short_strike - long_strike)
    credit = bid_price(short_leg) - ask_price(long_leg)

    if credit <= 0 or spread_width <= credit:
        return None

    credit_pct_of_width = credit / spread_width
    if credit_pct_of_width < config.min_credit_pct_of_width:
        return None

    short_distance_pct = (underlying_price - short_strike) / underlying_price if "Put" in strategy else (short_strike - underlying_price) / underlying_price
    if short_distance_pct < config.min_short_distance_pct:
        return None

    max_loss = spread_width - credit
    return_on_risk = credit / max_loss
    avg_open_interest = (_safe_int(short_leg.get("openInterest")) + _safe_int(long_leg.get("openInterest"))) / 2
    bid_ask_spread = (ask_price(short_leg) - bid_price(short_leg) + ask_price(long_leg) - bid_price(long_leg))

    dte = days_to_expiry(expiry)
    iv_for_calc = iv_decimal if iv_decimal > 0 else max(0.01, india_vix / 100.0)
    option_type = "PUT" if "Put" in strategy else "CALL"
    prob_of_profit = probability_otm(underlying_price, short_strike, iv_for_calc, dte, option_type)

    return {
        "Symbol": symbol,
        "Strategy": strategy,
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": short_leg["strikePrice"],
        "Buy Leg (Strike)": long_leg["strikePrice"],
        "Credit": round(credit, 2),
        "Net Debit": 0,
        "Spread Width": round(spread_width, 2),
        "Max Profit": round(credit, 2),
        "Max Loss": round(max_loss, 2),
        "R:R Ratio": round(return_on_risk, 2),
        "Return on Risk": round(return_on_risk, 2),
        "Short Strike Distance %": round(short_distance_pct * 100, 2),
        "Expected Move %": round(expected_move_pct(iv_for_calc, expiry) * 100, 2),
        "Probability of Profit": round(prob_of_profit, 3),
        "Avg OI": round(avg_open_interest),
        "Bid-Ask Spread": round(bid_ask_spread, 2),
        "Score": score_credit_spread(
            return_on_risk, prob_of_profit,
            avg_open_interest, max_open_interest, bid_ask_spread, credit_pct_of_width
        ),
    }


def build_short_iron_butterfly_opportunity(
    symbol: str, pcr: float, india_vix: float, underlying_price: float,
    atm_call: dict[str, Any], atm_put: dict[str, Any],
    otm_call: dict[str, Any], otm_put: dict[str, Any],
    max_open_interest: int, expiry: str, iv_decimal: float
) -> dict[str, Any] | None:
    atm_strike = float(atm_call["strikePrice"])
    otm_call_strike = float(otm_call["strikePrice"])
    otm_put_strike = float(otm_put["strikePrice"])

    call_wing_width = otm_call_strike - atm_strike
    put_wing_width = atm_strike - otm_put_strike

    if call_wing_width <= 0 or put_wing_width <= 0 or call_wing_width != put_wing_width:
        return None

    wing_width = call_wing_width
    credit = (bid_price(atm_call) + bid_price(atm_put)) - (ask_price(otm_call) + ask_price(otm_put))
    max_loss = wing_width - credit

    if credit <= 0 or max_loss <= 0:
        return None

    return_on_risk = credit / max_loss
    upper_be = atm_strike + credit
    lower_be = atm_strike - credit
    dte = days_to_expiry(expiry)

    prob_above_lower = probability_itm(underlying_price, lower_be, iv_decimal, dte, "CALL")
    prob_below_upper = probability_otm(underlying_price, upper_be, iv_decimal, dte, "CALL")
    prob_of_profit = max(0.0, min(1.0, prob_above_lower + prob_below_upper - 1.0))

    avg_oi = (
        _safe_int(atm_call.get("openInterest")) + _safe_int(atm_put.get("openInterest")) +
        _safe_int(otm_call.get("openInterest")) + _safe_int(otm_put.get("openInterest"))
    ) / 4
    bid_ask = (
        ask_price(atm_call) - bid_price(atm_call) + ask_price(atm_put) - bid_price(atm_put) +
        ask_price(otm_call) - bid_price(otm_call) + ask_price(otm_put) - bid_price(otm_put)
    )

    pop_score = min(35.0, prob_of_profit / 0.65 * 35.0)
    return_score = min(30.0, return_on_risk / 0.5 * 30.0)
    oi_score = (avg_oi / max_open_interest * 20.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 15.0 - min(15.0, bid_ask * 2.0))
    score = round(pop_score + return_score + oi_score + tightness_score, 2)

    return {
        "Symbol": symbol,
        "Strategy": "Short Iron Butterfly",
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": f"{atm_strike} CE + {atm_strike} PE",
        "Buy Leg (Strike)": f"{otm_put_strike} PE / {otm_call_strike} CE",
        "Credit": round(credit, 2),
        "Net Debit": 0,
        "Spread Width": round(wing_width, 2),
        "Max Profit": round(credit, 2),
        "Max Loss": round(max_loss, 2),
        "R:R Ratio": round(return_on_risk, 2),
        "Return on Risk": round(return_on_risk, 2),
        "Breakeven": f"{round(upper_be, 2)} / {round(lower_be, 2)}",
        "Probability of Profit": round(prob_of_profit, 3),
        "Expected Move %": round(expected_move_pct(iv_decimal, expiry) * 100, 2),
        "Avg OI": round(avg_oi),
        "Bid-Ask Spread": round(bid_ask, 2),
        "Score": score,
    }


def build_short_straddle_opportunity(
    symbol: str, pcr: float, india_vix: float, underlying_price: float,
    atm_call: dict[str, Any], atm_put: dict[str, Any],
    max_open_interest: int, expiry: str, iv_decimal: float,
) -> dict[str, Any] | None:
    call_strike = float(atm_call["strikePrice"])
    put_strike = float(atm_put["strikePrice"])
    if call_strike != put_strike:
        return None
    credit = bid_price(atm_call) + bid_price(atm_put)
    if credit <= 0:
        return None
    upper_be = call_strike + credit
    lower_be = put_strike - credit
    dte = days_to_expiry(expiry)
    prob_above_lower = probability_itm(underlying_price, lower_be, iv_decimal, dte, "CALL")
    prob_below_upper = probability_otm(underlying_price, upper_be, iv_decimal, dte, "CALL")
    prob_of_profit = max(0.0, min(1.0, prob_above_lower + prob_below_upper - 1.0))

    avg_oi = (_safe_int(atm_call.get("openInterest")) + _safe_int(atm_put.get("openInterest"))) / 2
    bid_ask = (ask_price(atm_call) - bid_price(atm_call) + ask_price(atm_put) - bid_price(atm_put))
    premium_pct = credit / underlying_price
    pop_score = min(35.0, prob_of_profit / 0.6 * 35.0)
    premium_score = min(25.0, premium_pct * 500)
    oi_score = (avg_oi / max_open_interest * 25.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 15.0 - min(15.0, bid_ask * 2.0))
    score = round(pop_score + premium_score + oi_score + tightness_score, 2)

    return {
        "Symbol": symbol,
        "Strategy": "Short Straddle",
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": f"{call_strike} CE + {put_strike} PE",
        "Buy Leg (Strike)": None,
        "Credit": round(credit, 2),
        "Net Debit": 0,
        "Spread Width": None,
        "Max Profit": round(credit, 2),
        "Max Loss": "Unlimited",
        "R:R Ratio": None,
        "Return on Risk": None,
        "Breakeven": f"{round(upper_be, 2)} / {round(lower_be, 2)}",
        "Probability of Profit": round(prob_of_profit, 3),
        "Expected Move %": round(expected_move_pct(iv_decimal, expiry) * 100, 2),
        "Avg OI": round(avg_oi),
        "Bid-Ask Spread": round(bid_ask, 2),
        "Score": score,
    }


def build_long_strangle_opportunity(
    symbol: str, pcr: float, india_vix: float, underlying_price: float,
    otm_call: dict[str, Any], otm_put: dict[str, Any],
    max_open_interest: int, expiry: str, iv_decimal: float,
) -> dict[str, Any] | None:
    call_strike = float(otm_call["strikePrice"])
    put_strike = float(otm_put["strikePrice"])
    if call_strike <= underlying_price or put_strike >= underlying_price:
        return None
    if call_strike <= put_strike:
        return None
    call_cost = ask_price(otm_call)
    put_cost = ask_price(otm_put)
    total_debit = call_cost + put_cost
    if total_debit <= 0:
        return None
    upper_be = call_strike + total_debit
    lower_be = put_strike - total_debit
    dte = days_to_expiry(expiry)
    expected_move = expected_move_pct(iv_decimal, expiry) * underlying_price
    prob_up = probability_itm(underlying_price, upper_be, iv_decimal, dte, "CALL")
    prob_down = probability_itm(underlying_price, lower_be, iv_decimal, dte, "PUT")
    prob_of_profit = min(1.0, prob_up + prob_down)

    avg_oi = (_safe_int(otm_call.get("openInterest")) + _safe_int(otm_put.get("openInterest"))) / 2
    bid_ask = (ask_price(otm_call) - bid_price(otm_call) + ask_price(otm_put) - bid_price(otm_put))
    move_efficiency = min(1.0, expected_move / total_debit) if total_debit > 0 else 0.0
    pop_score = min(35.0, prob_of_profit / 0.5 * 35.0)
    move_score = min(30.0, move_efficiency * 30.0)
    oi_score = (avg_oi / max_open_interest * 20.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 15.0 - min(15.0, bid_ask * 2.0))
    score = round(pop_score + move_score + oi_score + tightness_score, 2)

    return {
        "Symbol": symbol,
        "Strategy": "Long Strangle",
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": None,
        "Buy Leg (Strike)": f"{call_strike} CE + {put_strike} PE",
        "Net Debit": round(total_debit, 2),
        "Credit": 0,
        "Spread Width": round(call_strike - put_strike, 2),
        "Max Profit": "Unlimited",
        "Max Loss": round(total_debit, 2),
        "R:R Ratio": None,
        "Breakeven": f"{round(upper_be, 2)} / {round(lower_be, 2)}",
        "Probability of Profit": round(prob_of_profit, 3),
        "Expected Move %": round(expected_move_pct(iv_decimal, expiry) * 100, 2),
        "Avg OI": round(avg_oi),
        "Bid-Ask Spread": round(bid_ask, 2),
        "Score": score,
    }


def build_iron_condor_opportunity(
    symbol: str, pcr: float, india_vix: float, underlying_price: float,
    put_spread: dict[str, Any], call_spread: dict[str, Any], expiry: str
) -> dict[str, Any] | None:
    put_width = float(put_spread["Spread Width"])
    call_width = float(call_spread["Spread Width"])
    if put_width != call_width:
        return None

    credit = float(put_spread["Credit"]) + float(call_spread["Credit"])
    max_loss = put_width - credit
    if credit <= 0 or max_loss <= 0:
        return None

    avg_score = (float(put_spread["Score"]) + float(call_spread["Score"])) / 2
    return_on_risk = credit / max_loss
    distance_pct = min(float(put_spread["Short Strike Distance %"]), float(call_spread["Short Strike Distance %"]))
    pop_put = float(put_spread.get("Probability of Profit", 0.5))
    pop_call = float(call_spread.get("Probability of Profit", 0.5))
    prob_of_profit = max(0.0, min(1.0, pop_put + pop_call - 1.0))

    return {
        "Symbol": symbol,
        "Strategy": "Iron Condor",
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": f"{put_spread['Sell Leg (Strike)']} PE / {call_spread['Sell Leg (Strike)']} CE",
        "Buy Leg (Strike)": f"{put_spread['Buy Leg (Strike)']} PE / {call_spread['Buy Leg (Strike)']} CE",
        "Credit": round(credit, 2),
        "Net Debit": 0,
        "Spread Width": round(put_width, 2),
        "Max Profit": round(credit, 2),
        "Max Loss": round(max_loss, 2),
        "R:R Ratio": round(return_on_risk, 2),
        "Return on Risk": round(return_on_risk, 2),
        "Short Strike Distance %": round(distance_pct, 2),
        "Expected Move %": max(put_spread.get("Expected Move %", 0), call_spread.get("Expected Move %", 0)),
        "Probability of Profit": round(prob_of_profit, 3),
        "Avg OI": round((float(put_spread["Avg OI"]) + float(call_spread["Avg OI"])) / 2),
        "Bid-Ask Spread": round(float(put_spread["Bid-Ask Spread"]) + float(call_spread["Bid-Ask Spread"]), 2),
        "Score": round(min(100.0, avg_score + min(10.0, return_on_risk * 12.0)), 2),
    }


def near_underlying(option: dict[str, Any], underlying_price: float, window_pct: float) -> bool:
    strike = float(option.get("strikePrice") or 0)
    return underlying_price * (1 - window_pct) <= strike <= underlying_price * (1 + window_pct)


# ---------------------------------------------------------------------------
# Core Analysis Routines
# ---------------------------------------------------------------------------

def _collect_tradeable(
    records: list[dict[str, Any]], key: str, config: ScannerConfig,
    underlying_price: float, window_pct: float,
) -> list[dict[str, Any]]:
    return sorted(
        [
            row[key] for row in records
            if key in row
            and option_is_tradeable(row[key], config)
            and near_underlying(row[key], underlying_price, window_pct)
        ],
        key=lambda opt: float(opt["strikePrice"]),
    )


def _build_put_credit_spreads(
    symbol: str, context: MarketContext, config: ScannerConfig,
    india_vix: float, iv_decimal: float, puts: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results = []
    for short_idx, short_put in enumerate(puts):
        if float(short_put["strikePrice"]) >= context.underlying_price:
            continue
        for long_put in puts[max(0, short_idx - config.max_legs_apart):short_idx]:
            opp = build_credit_spread_opportunity(
                symbol, "Bull Put Credit Spread", context.pcr, india_vix,
                context.underlying_price, short_put, long_put,
                context.max_open_interest, context.expiry, config, iv_decimal,
            )
            if opp:
                opp = add_validation_fields(opp, context, config)
            if opp:
                results.append(opp)
    return results


def _build_call_credit_spreads(
    symbol: str, context: MarketContext, config: ScannerConfig,
    india_vix: float, iv_decimal: float, calls: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results = []
    for short_idx, short_call in enumerate(calls):
        if float(short_call["strikePrice"]) <= context.underlying_price:
            continue
        for long_call in calls[short_idx + 1:min(len(calls), short_idx + config.max_legs_apart + 1)]:
            opp = build_credit_spread_opportunity(
                symbol, "Bear Call Credit Spread", context.pcr, india_vix,
                context.underlying_price, short_call, long_call,
                context.max_open_interest, context.expiry, config, iv_decimal,
            )
            if opp:
                opp = add_validation_fields(opp, context, config)
            if opp:
                results.append(opp)
    return results


def _build_debit_spreads(
    symbol: str, strategy: str, context: MarketContext, config: ScannerConfig,
    india_vix: float, iv_decimal: float, options: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results = []
    for buy_idx, buy_leg in enumerate(options):
        for sell_leg in options[buy_idx + 1:min(len(options), buy_idx + config.max_legs_apart + 1)]:
            opp = build_opportunity(
                symbol, strategy, context.pcr, india_vix,
                context.underlying_price, buy_leg, sell_leg,
                context.max_open_interest, context.expiry, iv_decimal,
            )
            if opp:
                opp = add_validation_fields(opp, context, config)
            if opp:
                results.append(opp)
    return results


def _best_iron_condors(
    symbol: str, context: MarketContext, config: ScannerConfig,
    india_vix: float, put_spreads: list[dict[str, Any]], call_spreads: list[dict[str, Any]],
    top_n: int,
) -> list[dict[str, Any]]:
    condors = []
    for ps in put_spreads[:5]:
        for cs in call_spreads[:5]:
            if float(ps["Spread Width"]) != float(cs["Spread Width"]):
                continue
            condor = build_iron_condor_opportunity(
                symbol, context.pcr, india_vix, context.underlying_price,
                ps, cs, context.expiry,
            )
            if condor:
                condor = add_validation_fields(condor, context, config)
            if condor:
                condors.append(condor)
    condors.sort(key=lambda x: x["Score"], reverse=True)
    return condors[:max(1, top_n)]


def analyze_symbol(
    session: Any, symbol: str, india_vix: float, config: ScannerConfig,
    chain_type: str = "auto", expiry: str | None = None,
    trend_map: dict[str, str] | None = None, event_map: dict[str, str] | None = None,
    top_n: int = 1,
    strategy_mode: str = "both",
) -> list[dict[str, Any]]:
    data = fetch_option_chain(session, symbol, chain_type, expiry)
    if not data or "records" not in data:
        return []
    context = build_market_context(data, symbol, expiry, trend_map or {}, event_map or {}, config)
    if not context:
        return []

    sym = symbol.upper()
    iv = effective_iv(context, india_vix)
    trend = context.trend.lower()

    calls = _collect_tradeable(context.records, "CE", config, context.underlying_price, config.atm_window_pct)
    puts  = _collect_tradeable(context.records, "PE", config, context.underlying_price, config.atm_window_pct)

    candidates: list[dict[str, Any]] = []

    # ── BULLISH ──────────────────────────────────────────────────────────────
    if trend in ("bullish", "unknown"):
        if strategy_mode in ("buying", "both"):
            bull_calls = sorted(calls, key=lambda o: float(o["strikePrice"]))
            candidates.extend(_build_debit_spreads(sym, "Bull Call Spread", context, config, india_vix, iv, bull_calls))

        if strategy_mode in ("selling", "both"):
            candidates.extend(_build_put_credit_spreads(sym, context, config, india_vix, iv, puts))

    # ── BEARISH ──────────────────────────────────────────────────────────────
    if trend in ("bearish", "unknown"):
        if strategy_mode in ("buying", "both"):
            bear_puts = sorted(puts, key=lambda o: float(o["strikePrice"]), reverse=True)
            candidates.extend(_build_debit_spreads(sym, "Bear Put Spread", context, config, india_vix, iv, bear_puts))

        if strategy_mode in ("selling", "both"):
            candidates.extend(_build_call_credit_spreads(sym, context, config, india_vix, iv, calls))

    # ── SIDEWAYS / NEUTRAL ────────────────────────────────────────────────────
    if trend in ("sideways", "neutral", "unknown"):
        if calls and puts:
            atm_call = min(calls, key=lambda o: abs(float(o["strikePrice"]) - context.underlying_price))
            atm_put  = min(puts,  key=lambda o: abs(float(o["strikePrice"]) - context.underlying_price))
            otm_calls = [o for o in calls if float(o["strikePrice"]) > context.underlying_price]
            otm_puts  = [o for o in puts  if float(o["strikePrice"]) < context.underlying_price]

            if strategy_mode in ("selling", "both"):
                # 1. Iron Condor (Defined Risk Net Credit)
                put_spreads  = _build_put_credit_spreads(sym, context, config, india_vix, iv, puts)
                call_spreads = _build_call_credit_spreads(sym, context, config, india_vix, iv, calls)
                put_spreads.sort(key=lambda x: x["Score"], reverse=True)
                call_spreads.sort(key=lambda x: x["Score"], reverse=True)
                if put_spreads and call_spreads:
                    candidates.extend(_best_iron_condors(sym, context, config, india_vix, put_spreads, call_spreads, top_n))

                # 2. Short Iron Butterfly (Defined Risk Net Credit, Tight ATM Setup)
                if otm_calls and otm_puts:
                    bf_pool = []
                    for oc in otm_calls[:config.max_legs_apart]:
                        for op in otm_puts[-config.max_legs_apart:]:
                            opp = build_short_iron_butterfly_opportunity(
                                sym, context.pcr, india_vix, context.underlying_price,
                                atm_call, atm_put, oc, op, context.max_open_interest, context.expiry, iv
                            )
                            if opp:
                                opp = add_validation_fields(opp, context, config)
                            if opp:
                                bf_pool.append(opp)
                    bf_pool.sort(key=lambda x: x["Score"], reverse=True)
                    candidates.extend(bf_pool[:max(1, top_n)])

                # 3. Short Straddle (Selling Undefined Risk Volatility)
                opp = build_short_straddle_opportunity(
                    sym, context.pcr, india_vix, context.underlying_price,
                    atm_call, atm_put, context.max_open_interest, context.expiry, iv,
                )
                if opp:
                    opp = add_validation_fields(opp, context, config)
                if opp:
                    candidates.append(opp)

            if strategy_mode in ("buying", "both") and trend == "sideways":
                # Long Strangle: buy OTM call + OTM put (only when compression confirmed)
                strangle_pool = []
                for oc in otm_calls[:config.max_legs_apart]:
                    for op in otm_puts[-config.max_legs_apart:]:
                        opp = build_long_strangle_opportunity(
                            sym, context.pcr, india_vix, context.underlying_price,
                            oc, op, context.max_open_interest, context.expiry, iv,
                        )
                        if opp:
                            opp = add_validation_fields(opp, context, config)
                        if opp:
                            strangle_pool.append(opp)
                strangle_pool.sort(key=lambda x: x["Score"], reverse=True)
                candidates.extend(strangle_pool[:max(1, top_n)])

    candidates.sort(key=lambda x: x["Score"], reverse=True)
    return candidates[:max(1, top_n)]


def write_results(results: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(results).sort_values(["Score", "R:R Ratio"], ascending=False)

    if output_path.suffix.lower() == ".csv":
        df.to_csv(output_path, index=False)
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    headers = list(df.columns)
    has_validation_col = "Validation Pass" in headers

    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Opportunities"

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    bold = Font(bold=True)

    def write_sheet(ws, frame: pd.DataFrame, highlight_failed: bool) -> None:
        for row in dataframe_to_rows(frame, index=False, header=True):
            ws.append(row)
        for cell in ws[1]:
            cell.font = bold
        if highlight_failed and has_validation_col:
            val_col_idx = headers.index("Validation Pass") + 1
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=val_col_idx).value is False:
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = red_fill
                        cell.font = red_font
        ws.freeze_panes = "A2"
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            values = [str(header)] + [str(v) for v in frame[header].tolist()]
            max_len = max(len(v) for v in values)
            ws.column_dimensions[col_letter].width = min(28, max(10, max_len + 2))

    write_sheet(ws_all, df, highlight_failed=True)

    if has_validation_col:
        strict_df = df[df["Validation Pass"] == True].sort_values(
            ["Score", "R:R Ratio"], ascending=False
        )
        ws_strict = wb.create_sheet("Strict (enforce-validations)")
        write_sheet(ws_strict, strict_df, highlight_failed=False)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Integrated Scanner: Technical Trend + NSE Option Chain Spread Analysis.")
    parser.add_argument("--symbols", nargs="+", default=list(DEFAULT_SYMBOLS))
    parser.add_argument("--output", default="Combined_Option_Spread_Analysis.xlsx")
    parser.add_argument("--min-oi", type=int, default=100)
    parser.add_argument("--min-bid", type=float, default=0.05)
    parser.add_argument("--atm-window-pct", type=float, default=0.12)
    parser.add_argument("--max-legs-apart", type=int, default=8)
    parser.add_argument("--expiry", help="NSE expiry format e.g. 25-Aug-2026")
    parser.add_argument("--chain-type", choices=("auto", "equity", "indices"), default="auto")
    parser.add_argument("--strategy-mode", choices=("buying", "selling", "both"), default="both")
    parser.add_argument("--min-credit-pct-of-width", type=float, default=0.12)
    parser.add_argument("--min-short-distance-pct", type=float, default=0.01)
    parser.add_argument("--max-bid-ask-spread-pct", type=float, default=0.25)
    parser.add_argument("--min-volume", type=int, default=10)
    parser.add_argument("--top-n", type=int, default=1)
    parser.add_argument("--lot-size", type=int, default=1)
    parser.add_argument("--max-margin-per-trade", type=float)
    parser.add_argument("--trend-file", help="CSV with Symbol,Trend columns.")
    parser.add_argument("--event-file", help="CSV with Symbol,EventRisk columns.")
    parser.add_argument("--enforce-validations", action="store_true")
    parser.add_argument("--trend-debug", action="store_true")
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--cookie-header")
    parser.add_argument("--cookie-file")
    parser.add_argument("--browser-impersonation", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO if args.verbose else logging.WARNING, format="%(levelname)s: %(message)s")

    config = ScannerConfig(
        min_bid=args.min_bid,
        min_open_interest=args.min_oi,
        max_legs_apart=args.max_legs_apart,
        atm_window_pct=args.atm_window_pct,
        min_credit_pct_of_width=args.min_credit_pct_of_width,
        min_short_distance_pct=args.min_short_distance_pct,
        max_bid_ask_spread_pct=args.max_bid_ask_spread_pct,
        min_volume=args.min_volume,
        max_margin_per_trade=args.max_margin_per_trade,
        lot_size=args.lot_size,
        enforce_validations=args.enforce_validations,
    )

    trend_map = load_symbol_map(args.trend_file, "Trend")
    event_map = load_symbol_map(args.event_file, "EventRisk")
    cookie_header = load_cookie_header(args.cookie_header, args.cookie_file)

    try:
        session = create_nse_session(cookie_header, args.browser_impersonation)
    except Exception as exc:
        raise SystemExit(f"Unable to initialize NSE session: {exc}") from exc

    india_vix = fetch_india_vix(session, config.fallback_vix)
    print(f"Current India VIX: {india_vix:.2f}")

    trend_diagnostics: list[dict[str, Any]] = []

    results: list[dict[str, Any]] = []
    for symbol in args.symbols:
        print(f"Analyzing {symbol.upper()}...")

        if args.trend_debug:
            _, diag = detect_market_trend_detailed(symbol.upper(), lookback_period=config.lookback_period)
            trend_diagnostics.append(diag)

        results.extend(
            analyze_symbol(
                session, symbol, india_vix, config,
                chain_type=args.chain_type, expiry=args.expiry,
                trend_map=trend_map, event_map=event_map,
                top_n=args.top_n, strategy_mode=args.strategy_mode,
            )
        )

    if args.trend_debug and trend_diagnostics:
        debug_path = Path(args.output).with_name(Path(args.output).stem + "_trend_debug.csv")
        pd.DataFrame(trend_diagnostics).to_csv(debug_path, index=False)
        print(f"Trend diagnostics saved to '{debug_path}'.")

    if not results:
        print("No qualifying spread opportunities found.")
        return

    output_path = Path(args.output)
    write_results(results, output_path)
    print(f"Analysis completed. Output saved to '{output_path}'.")


if __name__ == "__main__":
    main()