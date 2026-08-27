import argparse
import logging
import math
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import requests

try:
    import yfinance as yf
except ImportError:
    yf = None

from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from curl_cffi import requests as curl_requests
except ImportError:
    curl_requests = None


NSE_HOME_URL = "https://www.nseindia.com"
INDICES_URL = "https://www.nseindia.com/api/allIndices"
OPTION_CHAIN_CONTRACT_INFO_URL = "https://www.nseindia.com/api/option-chain-contract-info"
OPTION_CHAIN_URL = "https://www.nseindia.com/api/option-chain-v3"

DEFAULT_SYMBOLS = (
    "BIOCON",
    "ADANIPORTS",
    "INFY",
    "HDFCBANK",
    "ICICIBANK",
    "SBIN",
    "BHARTIARTL",
    "BHEL",
    "HINDPETRO",
    "IOC"
)

INDEX_SYMBOLS = {"NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "NIFTYNXT50"}

# Yahoo Finance identifiers for NSE index option underlyings (NIFTY.NS / BANKNIFTY.NS are invalid).
INDEX_YAHOO_ALIASES = {
    "NIFTY": "^NSEI",
    "NIFTY50": "^NSEI",
    "BANKNIFTY": "^NSEBANK",
    "FINNIFTY": "NIFTY_FIN_SERVICE.NS",
    "MIDCPNIFTY": "NIFTY_MID_SELECT.NS",
}


def yahoo_history_ticker(symbol: str) -> str:
    """Map friendly NSE names / indices to Yahoo Finance history tickers."""
    cleaned = symbol.strip().upper()
    if cleaned in INDEX_YAHOO_ALIASES:
        return INDEX_YAHOO_ALIASES[cleaned]
    if "." in cleaned or cleaned.startswith("^"):
        return cleaned
    return f"{cleaned}.NS"


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
    "DNT": "1",
    "Referer": NSE_HOME_URL,
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Upgrade-Insecure-Requests": "1",
}

# Fast O(1) strategy directional mapping for confluence verification
STRATEGY_DIRECTION_MAP = {
    "Bull Call Spread": "BULLISH",
    "Bull Put Credit Spread": "BULLISH",
    "Bear Put Spread": "BEARISH",
    "Bear Call Credit Spread": "BEARISH",
    "Iron Condor": "NEUTRAL",
    "Long Strangle": "NEUTRAL",
    "Short Straddle": "NEUTRAL",
}

LOGGER = logging.getLogger("combined_option_scanner")


@dataclass(frozen=True)
class ScannerConfig:
    min_bid: float = 0.05
    min_open_interest: int = 100
    max_legs_apart: int = 8
    atm_window_pct: float = 0.12
    neutral_pcr_low: float = 0.7
    neutral_pcr_high: float = 1.3
    fallback_vix: float = 15.0
    min_credit_pct_of_width: float = 0.12
    min_short_distance_pct: float = 0.01
    max_bid_ask_spread_pct: float = 0.25
    min_volume: int = 10
    max_margin_per_trade: float | None = None
    lot_size: int = 1
    enforce_validations: bool = False
    lookback_period: str = "1y"


@dataclass(frozen=True)
class MarketContext:
    symbol: str
    records: list[dict[str, Any]]
    underlying_price: float
    pcr: float
    max_open_interest: int
    expiry: str
    trend: str = "unknown"
    event_risk: str = "unknown"
    atm_iv: float = 0.0  # decimal (e.g. 0.18 for 18%), derived from the symbol's own chain


def effective_iv(context: "MarketContext", india_vix: float) -> float:
    """Symbol-specific IV is far more accurate for expected-move math than a blanket
    index VIX applied to every stock. Fall back to India VIX only when the chain
    doesn't expose usable implied vol (e.g. illiquid strikes near the money)."""
    if context.atm_iv and context.atm_iv > 0:
        return context.atm_iv
    return max(0.01, india_vix / 100.0)


# ---------------------------------------------------------------------------
# Technical Trend Scanners Logic (Bullish, Bearish, Rangebound/Strangle)
# ---------------------------------------------------------------------------

def calculate_technical_indicators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # EMAs
    df["EMA9"] = df["Close"].ewm(span=9, adjust=False).mean()
    df["EMA18"] = df["Close"].ewm(span=18, adjust=False).mean()
    df["EMA50"] = df["Close"].ewm(span=50, adjust=False).mean()
    df["EMA200"] = df["Close"].ewm(span=200, adjust=False).mean()

    # CCI (14 & 20)
    tp = (df["High"] + df["Low"] + df["Close"]) / 3
    sma_tp14 = tp.rolling(window=14).mean()
    mad14 = tp.rolling(window=14).apply(lambda x: np.abs(x - x.mean()).mean(), raw=True)
    df["CCI14"] = (tp - sma_tp14) / (0.015 * mad14)

    sma_tp20 = tp.rolling(window=20).mean()
    mad20 = tp.rolling(window=20).apply(lambda x: np.abs(x - x.mean()).mean(), raw=True)
    df["CCI20"] = (tp - sma_tp20) / (0.015 * mad20)

    # ADX (14)
    high_diff = df["High"].diff()
    low_diff = -df["Low"].diff()
    pos_dm = np.where((high_diff > low_diff) & (high_diff > 0), high_diff, 0.0)
    neg_dm = np.where((low_diff > high_diff) & (low_diff > 0), low_diff, 0.0)

    tr1 = df["High"] - df["Low"]
    tr2 = np.abs(df["High"] - df["Close"].shift(1))
    tr3 = np.abs(df["Low"] - df["Close"].shift(1))
    tr = pd.DataFrame({"tr1": tr1, "tr2": tr2, "tr3": tr3}).max(axis=1)

    atr = tr.rolling(window=14).mean()
    pos_di = 100 * (pd.Series(pos_dm, index=df.index).rolling(window=14).mean() / atr)
    neg_di = 100 * (pd.Series(neg_dm, index=df.index).rolling(window=14).mean() / atr)
    dx = 100 * (np.abs(pos_di - neg_di) / (pos_di + neg_di))
    df["ADX"] = dx.rolling(window=14).mean()

    # Fibonacci S1
    prev_high = df["High"].shift(1)
    prev_low = df["Low"].shift(1)
    prev_close = df["Close"].shift(1)
    pivot = (prev_high + prev_low + prev_close) / 3
    df["Fib_S1"] = pivot - 0.382 * (prev_high - prev_low)

    return df


def detect_market_trend(symbol: str, lookback_period: str = "1y") -> str:
    """Fetches market price data and derives technical trend: bullish, bearish, sideways, or neutral."""
    trend, _ = detect_market_trend_detailed(symbol, lookback_period)
    return trend


def detect_market_trend_detailed(
    symbol: str, lookback_period: str = "1y"
) -> tuple[str, dict[str, Any]]:
    """Same logic as detect_market_trend, but also returns a diagnostics dict with the
    raw indicator values and each individual condition's pass/fail for the latest bar.

    This exists so a "neutral" result can be inspected rather than trusted blindly --
    each of bullish/bearish/sideways requires 4-7 conditions to align simultaneously
    (faithfully ported from bullishbiasnifty500.py / bearisbiasnifty500.py /
    rangeboundstocks.py), so it's expected that most symbols land on neutral on most
    days. This diagnostic makes it possible to tell that apart from a real data problem
    (bad ticker suffix, too-short history, NaN indicators, etc.)."""
    diag: dict[str, Any] = {"Symbol": symbol, "Data Rows": 0, "Fetch Error": ""}

    if yf is None:
        LOGGER.warning("yfinance package not installed. Skipping dynamic trend detection.")
        diag["Fetch Error"] = "yfinance not installed"
        return "unknown", diag

    formatted_ticker = yahoo_history_ticker(symbol)
    diag["Ticker Used"] = formatted_ticker
    try:
        ticker = yf.Ticker(formatted_ticker)
        df = ticker.history(period=lookback_period, interval="1d")
        diag["Data Rows"] = len(df)

        if df.empty or len(df) < 50:
            diag["Fetch Error"] = f"Insufficient history ({len(df)} rows, need >=50)"
            return "unknown", diag

        df = calculate_technical_indicators(df)
        curr = df.iloc[-1]
        prev = df.iloc[-2]

        diag.update({
            "Close": round(curr["Close"], 2),
            "EMA9": round(curr["EMA9"], 2),
            "EMA18": round(curr["EMA18"], 2),
            "EMA50": round(curr["EMA50"], 2),
            "EMA200": round(curr["EMA200"], 2) if len(df) >= 200 else None,
            "CCI14": round(curr["CCI14"], 2) if pd.notna(curr["CCI14"]) else None,
            "CCI20": round(curr["CCI20"], 2) if pd.notna(curr["CCI20"]) else None,
            "ADX": round(curr["ADX"], 2) if pd.notna(curr["ADX"]) else None,
            "Fib_S1": round(curr["Fib_S1"], 2) if pd.notna(curr["Fib_S1"]) else None,
        })

        # 1. Bullish Setup
        bullish_ready = len(df) >= 200
        diag["Bullish: enough history (>=200d)"] = bullish_ready
        if bullish_ready:
            price_above_emas = (
                curr["Close"] > curr["EMA9"]
                and curr["EMA9"] > curr["EMA18"]
                and curr["EMA18"] > curr["EMA50"]
                and curr["EMA50"] > curr["EMA200"]
            )
            ema200_sloping_up = curr["EMA200"] >= df.iloc[-5]["EMA200"]
            cci_trigger = (prev["CCI20"] <= 100 and curr["CCI20"] > 100) or (curr["CCI20"] > 100)
            body_size = abs(curr["Close"] - curr["Open"])
            candle_range = curr["High"] - curr["Low"]
            is_bullish_candle = (curr["Close"] > curr["Open"]) and (
                body_size / candle_range >= 0.5 if candle_range > 0 else True
            )
            riding_ema9 = curr["Low"] >= curr["EMA9"] or curr["Close"] > curr["EMA9"]
            adx_pass_bull = curr["ADX"] >= 20.0

            diag["Bullish: EMA stack (P>9>18>50>200)"] = price_above_emas
            diag["Bullish: EMA200 sloping up"] = ema200_sloping_up
            diag["Bullish: CCI20 > 100 trigger"] = cci_trigger
            diag["Bullish: bullish candle body>=50%"] = is_bullish_candle
            diag["Bullish: riding EMA9"] = riding_ema9
            diag["Bullish: ADX >= 20"] = adx_pass_bull

            if (
                price_above_emas and ema200_sloping_up and cci_trigger
                and is_bullish_candle and riding_ema9 and adx_pass_bull
            ):
                diag["Result"] = "bullish"
                return "bullish", diag

        # 2. Bearish Setup
        ema_alignment_bearish = (
            curr["EMA50"] > curr["EMA18"]
            and curr["EMA18"] > curr["EMA9"]
            and curr["Close"] < curr["EMA9"]
        )
        adx_pass_bear = curr["ADX"] >= 20.0
        cci_pass_bear = curr["CCI14"] <= -100.0
        fib_pass = curr["Close"] < curr["Fib_S1"]

        diag["Bearish: ADX >= 20"] = adx_pass_bear
        diag["Bearish: CCI14 <= -100"] = cci_pass_bear
        diag["Bearish: EMA stack (50>18>9>Price)"] = ema_alignment_bearish
        diag["Bearish: Close < Fib S1"] = fib_pass

        if adx_pass_bear and cci_pass_bear and ema_alignment_bearish and fib_pass:
            diag["Result"] = "bearish"
            return "bearish", diag

        # 3. Rangebound / Strangle Compression Setup
        recent_df = df.iloc[-3:]
        adx_below_threshold = (recent_df["ADX"] < 15.0).any()
        cci_tight = -50.0 <= curr["CCI14"] <= 50.0

        ema_min = min(curr["EMA9"], curr["EMA18"], curr["EMA50"])
        ema_max = max(curr["EMA9"], curr["EMA18"], curr["EMA50"])
        ema_spread_pct = (ema_max - ema_min) / curr["Close"]
        ema_braid_pass = ema_spread_pct <= 0.012

        box_df = df.iloc[-6:]
        box_high = box_df["High"].max()
        box_low = box_df["Low"].min()
        box_range_pct = (box_high - box_low) / curr["Close"]
        box_pass = box_range_pct <= 0.035

        diag["Sideways: ADX < 15 (last 3d)"] = adx_below_threshold
        diag["Sideways: CCI14 in [-50,50]"] = cci_tight
        diag["Sideways: EMA braid <=1.2%"] = ema_braid_pass
        diag["Sideways: EMA braid spread %"] = round(ema_spread_pct * 100, 2)
        diag["Sideways: 6d box range <=3.5%"] = box_pass
        diag["Sideways: 6d box range %"] = round(box_range_pct * 100, 2)

        if adx_below_threshold and cci_tight and ema_braid_pass and box_pass:
            diag["Result"] = "sideways"
            return "sideways", diag

        diag["Result"] = "neutral"
        return "neutral", diag

    except Exception as exc:
        LOGGER.warning("Failed to calculate trend for %s: %s", symbol, exc)
        diag["Fetch Error"] = str(exc)
        return "unknown", diag


# ---------------------------------------------------------------------------
# NSE API & Cookie Handling
# ---------------------------------------------------------------------------

def load_cookie_header(cookie_header: str | None, cookie_file: str | None) -> str | None:
    if cookie_header:
        return cookie_header.strip()
    if cookie_file:
        return Path(cookie_file).read_text(encoding="utf-8").strip()
    return None


def create_nse_session(
    cookie_header: str | None = None,
    use_browser_impersonation: bool = False,
) -> Any:
    if use_browser_impersonation:
        if curl_requests is None:
            raise RuntimeError("curl_cffi is not installed. Run: pip install curl_cffi")
        session = curl_requests.Session(impersonate="chrome120")
    else:
        session = requests.Session()

    session.headers.update(HEADERS)
    if cookie_header:
        session.headers.update({"Cookie": cookie_header})

    if not use_browser_impersonation:
        retry = Retry(
            total=3, connect=3, read=3, backoff_factor=0.5,
            status_forcelist=(429, 500, 502, 503, 504), allowed_methods=("GET",)
        )
        session.mount("https://", HTTPAdapter(max_retries=retry))

    response = session.get(NSE_HOME_URL, timeout=10)
    response.raise_for_status()
    session.get(f"{NSE_HOME_URL}/option-chain", timeout=10)
    return session


def get_json(session: Any, url: str, **params: Any) -> dict[str, Any] | None:
    api_headers = {
        **HEADERS,
        "Accept": "application/json,text/plain,*/*",
        "Referer": f"{NSE_HOME_URL}/option-chain",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "Upgrade-Insecure-Requests": "0",
    }
    try:
        response = session.get(url, params=params or None, timeout=15, headers=api_headers)
        if response.status_code == 403:
            LOGGER.warning("NSE returned 403 for %s.", url)
            return None
        response.raise_for_status()
        return response.json()
    except Exception as exc:
        LOGGER.warning("Request failed for %s: %s", url, exc)
    return None


def fetch_india_vix(session: Any, fallback: float) -> float:
    payload = get_json(session, INDICES_URL)
    for index in (payload or {}).get("data", []):
        if index.get("index") == "INDIA VIX":
            try:
                return float(index["last"])
            except (KeyError, TypeError, ValueError):
                break
    return fallback


def fetch_contract_info(session: Any, symbol: str) -> dict[str, Any] | None:
    return get_json(session, OPTION_CHAIN_CONTRACT_INFO_URL, symbol=symbol.upper())


def detect_chain_type(symbol: str, chain_type: str) -> str:
    if chain_type != "auto":
        return "Indices" if chain_type == "indices" else "Equity"
    return "Indices" if symbol.upper() in INDEX_SYMBOLS else "Equity"


def fetch_option_chain(
    session: Any, symbol: str, chain_type: str, expiry: str | None = None
) -> dict[str, Any] | None:
    symbol = symbol.upper()
    contract_info = fetch_contract_info(session, symbol)
    expiry_dates = (contract_info or {}).get("expiryDates") or []
    selected_expiry = expiry or (expiry_dates[0] if expiry_dates else None)
    if not selected_expiry:
        return None

    payload = get_json(
        session,
        OPTION_CHAIN_URL,
        type=detect_chain_type(symbol, chain_type),
        symbol=symbol,
        expiry=selected_expiry,
    )
    if payload:
        payload["_selected_expiry"] = selected_expiry
    return payload


def latest_expiry_records(data: dict[str, Any]) -> list[dict[str, Any]]:
    return data.get("records", {}).get("data") or []


def load_symbol_map(path: str | None, value_column: str) -> dict[str, str]:
    if not path:
        return {}
    df = pd.read_csv(path)
    normalized_columns = {"".join(str(col).lower().split()): col for col in df.columns}
    symbol_column = normalized_columns.get("symbol")
    value_column_name = normalized_columns.get("".join(value_column.lower().split()))
    if not symbol_column or not value_column_name:
        raise ValueError(f"{path} must contain Symbol and {value_column} columns")
    return {
        str(row[symbol_column]).upper(): str(row[value_column_name]).strip()
        for _, row in df.iterrows()
        if pd.notna(row[symbol_column]) and pd.notna(row[value_column_name])
    }


def bid_price(option: dict[str, Any]) -> float:
    return float(option.get("bidprice") or option.get("buyPrice1") or 0)


def ask_price(option: dict[str, Any]) -> float:
    return float(option.get("askPrice") or option.get("sellPrice1") or 0)


def traded_volume(option: dict[str, Any]) -> int:
    return int(option.get("totalTradedVolume") or option.get("volume") or 0)


def bid_ask_spread_pct(option: dict[str, Any]) -> float:
    ask = ask_price(option)
    return max(0.0, ask - bid_price(option)) / ask if ask > 0 else 1.0


def option_is_tradeable(option: dict[str, Any], config: ScannerConfig) -> bool:
    return (
        bid_price(option) >= config.min_bid
        and ask_price(option) > 0
        and int(option.get("openInterest") or 0) >= config.min_open_interest
        and traded_volume(option) >= config.min_volume
        and bid_ask_spread_pct(option) <= config.max_bid_ask_spread_pct
    )


def build_market_context(
    data: dict[str, Any],
    symbol: str,
    expiry: str | None,
    trend_map: dict[str, str],
    event_map: dict[str, str],
    config: ScannerConfig,
) -> MarketContext | None:
    records = latest_expiry_records(data)
    if not records:
        return None

    underlying_price = float(data["records"].get("underlyingValue") or 0)
    if underlying_price <= 0:
        for row in records:
            underlying_price = float(
                row.get("CE", {}).get("underlyingValue")
                or row.get("PE", {}).get("underlyingValue")
                or 0
            )
            if underlying_price > 0:
                break
    if underlying_price <= 0:
        return None

    total_ce_oi = sum(int(row.get("CE", {}).get("openInterest") or 0) for row in records)
    total_pe_oi = sum(int(row.get("PE", {}).get("openInterest") or 0) for row in records)
    pcr = total_pe_oi / total_ce_oi if total_ce_oi > 0 else 1.0
    max_open_interest = max(
        [int(row.get("CE", {}).get("openInterest") or 0) for row in records]
        + [int(row.get("PE", {}).get("openInterest") or 0) for row in records]
        + [1]
    )

    symbol_key = symbol.upper()
    trend = trend_map.get(symbol_key, "").lower()

    if not trend or trend == "unknown":
        trend = detect_market_trend(symbol_key, lookback_period=config.lookback_period)

    # Symbol-specific ATM implied vol, averaged across CE/PE at the strike closest to
    # spot. This replaces the prior approach of using the blanket India VIX (an INDEX
    # vol) as the expected-move input for every individual stock, which systematically
    # mis-sized expected moves for stocks whose IV differs meaningfully from the index.
    atm_iv = 0.0
    closest_row = min(
        records,
        key=lambda row: abs(
            float((row.get("CE") or {}).get("strikePrice") or (row.get("PE") or {}).get("strikePrice") or 0)
            - underlying_price
        ),
    )
    ce_iv = float((closest_row.get("CE") or {}).get("impliedVolatility") or 0)
    pe_iv = float((closest_row.get("PE") or {}).get("impliedVolatility") or 0)
    ivs = [v for v in (ce_iv, pe_iv) if v > 0]
    if ivs:
        atm_iv = (sum(ivs) / len(ivs)) / 100.0

    return MarketContext(
        symbol=symbol_key,
        records=records,
        underlying_price=underlying_price,
        pcr=pcr,
        max_open_interest=max_open_interest,
        expiry=data.get("_selected_expiry") or expiry or "",
        trend=trend,
        event_risk=event_map.get(symbol_key, "unknown").lower(),
        atm_iv=atm_iv,
    )


# O(1) Dictionary Lookup for Trend Alignment
def trend_allows_strategy(strategy: str, trend: str) -> bool:
    if trend in ("unknown", "neutral", ""):
        return True

    required_direction = STRATEGY_DIRECTION_MAP.get(strategy)
    if not required_direction:
        return True

    trend_upper = trend.upper()
    if required_direction == "BULLISH":
        return trend_upper == "BULLISH"
    elif required_direction == "BEARISH":
        return trend_upper == "BEARISH"
    elif required_direction == "NEUTRAL":
        return trend_upper == "SIDEWAYS"

    return False


def pcr_bias(context: MarketContext, config: ScannerConfig) -> str:
    if context.pcr < config.neutral_pcr_low:
        return "bullish"
    if context.pcr > config.neutral_pcr_high:
        return "bearish"
    return "neutral"


# O(1) Dictionary Lookup for PCR Alignment
def pcr_allows_strategy(strategy: str, bias: str) -> bool:
    required_direction = STRATEGY_DIRECTION_MAP.get(strategy)
    if not required_direction:
        return True

    bias_upper = bias.upper()
    if required_direction == "BULLISH":
        return bias_upper == "BULLISH"
    elif required_direction == "BEARISH":
        return bias_upper == "BEARISH"
    elif required_direction == "NEUTRAL":
        return bias_upper == "NEUTRAL"

    return False


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def add_validation_fields(
    opportunity: dict[str, Any] | None, context: MarketContext, config: ScannerConfig
) -> dict[str, Any] | None:
    if opportunity is None:
        return None
    strategy = str(opportunity["Strategy"])
    bias = pcr_bias(context, config)

    # 1. Early Out Optimization: Check PCR Bias first
    bias_pass = pcr_allows_strategy(strategy, bias)
    if config.enforce_validations and not bias_pass:
        return None  # Skip rest of heavy calculations if PCR fails in enforced mode

    # 2. Additional Checks
    # Max Loss/Profit can be the string "Unlimited" for undefined-risk strategies
    # (Short Straddle = unlimited loss, Long Strangle/Straddle = unlimited profit).
    # Use a conservative SPAN-style margin proxy for those instead of crashing on
    # float("Unlimited") -- actual margin depends on the broker's real-time
    # SPAN+exposure calc, which isn't derivable from the option chain alone.
    raw_max_loss = opportunity.get("Max Loss")
    if isinstance(raw_max_loss, str) and raw_max_loss.strip().lower() == "unlimited":
        margin_per_unit = round(context.underlying_price * 0.20, 2)
    else:
        margin_per_unit = (
            _safe_float(raw_max_loss)
            or _safe_float(opportunity.get("Net Debit"))
            or _safe_float(opportunity.get("Credit"))
            or 0.0
        )
    estimated_margin = round(margin_per_unit * config.lot_size, 2)
    spread_width = opportunity.get("Spread Width") or 0
    liquidity_pass = (
        float(opportunity["Avg OI"]) >= config.min_open_interest
        and (
            float(spread_width) <= 0
            or float(opportunity["Bid-Ask Spread"]) <= float(spread_width) * config.max_bid_ask_spread_pct
        )
    )
    trend_pass = trend_allows_strategy(strategy, context.trend)
    event_pass = context.event_risk not in ("yes", "true", "high", "blocked", "avoid")
    margin_pass = (
        config.max_margin_per_trade is None or estimated_margin <= config.max_margin_per_trade
    )

    # Confluence Integrity: Both PCR bias AND technical trend (along with risk/margin checks) must pass
    validation_pass = bias_pass and trend_pass and liquidity_pass and event_pass and margin_pass

    # Reward genuine PCR + technical-trend confluence in the score itself, not just as a
    # pass/fail gate. Previously, unless --enforce-validations was passed, a trade that
    # fought the trend scored identically to one that agreed with it -- the two signals
    # were computed but never actually influenced ranking. Now every run's ranking
    # reflects confluence, while --enforce-validations remains available as a hard filter.
    strategy_direction = STRATEGY_DIRECTION_MAP.get(strategy)
    confluence_bonus = 0.0
    if strategy_direction:
        if bias.upper() == strategy_direction:
            confluence_bonus += 4.0
        trend_upper = context.trend.upper()
        trend_matches = (
            trend_upper == strategy_direction
            or (strategy_direction == "NEUTRAL" and trend_upper == "SIDEWAYS")
        )
        if trend_matches:
            confluence_bonus += 6.0

    opportunity["Score"] = round(min(100.0, float(opportunity["Score"]) + confluence_bonus), 2)

    opportunity.update(
        {
            "Trend": context.trend,
            "Trend Pass": trend_pass,
            "PCR Bias": bias,
            "PCR Bias Pass": bias_pass,
            "Confluence Bonus": confluence_bonus,
            "Event Risk": context.event_risk,
            "Event Risk Pass": event_pass,
            "Liquidity Pass": liquidity_pass,
            "Estimated Margin": estimated_margin,
            "Margin Pass": margin_pass,
            "Validation Pass": validation_pass,
        }
    )

    if config.enforce_validations and not validation_pass:
        return None

    return opportunity


# ---------------------------------------------------------------------------
# Scoring and Opportunity Builders
# ---------------------------------------------------------------------------

def score_spread(
    rr_ratio: float, cost_efficiency: float, avg_open_interest: float,
    max_open_interest: int, total_bid_ask_spread: float, prob_of_profit: float = 0.5
) -> float:
    # Weights re-balanced (out of 100) to make room for a real probability-of-profit
    # term, since R:R alone says nothing about how *likely* the trade is to work.
    rr_score = min(30.0, (rr_ratio / 3.0) * 30.0) if rr_ratio > 0 else 0.0
    cost_efficiency_score = max(0.0, (1.0 - cost_efficiency) * 20.0)
    pop_score = min(25.0, max(0.0, prob_of_profit) / 0.6 * 25.0)
    oi_score = (avg_open_interest / max_open_interest * 15.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 10.0 - min(10.0, total_bid_ask_spread * 2.0))
    return round(rr_score + cost_efficiency_score + pop_score + oi_score + tightness_score, 2)


def parse_nse_expiry(expiry: str) -> date | None:
    for fmt in ("%d-%b-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(expiry, fmt).date()
        except ValueError:
            continue
    return None


def days_to_expiry(expiry: str) -> int:
    expiry_date = parse_nse_expiry(expiry)
    return max(1, (expiry_date - date.today()).days) if expiry_date else 1


def expected_move_pct(iv_decimal: float, expiry: str) -> float:
    """iv_decimal should be the SYMBOL's own implied vol (decimal), not a blanket index VIX."""
    return max(0.005, iv_decimal * (days_to_expiry(expiry) / 365) ** 0.5)


def _norm_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def probability_otm(
    underlying_price: float, strike: float, iv_decimal: float, dte_days: int, option_type: str
) -> float:
    """Black-Scholes (zero-rate, zero-dividend) probability that an option finishes
    out-of-the-money at expiry. This is a standard proxy for probability of profit on
    a short option and is far more informative than a flat 'distance vs VIX' heuristic
    because it accounts for the strike's own implied vol and time to expiry directly."""
    if underlying_price <= 0 or strike <= 0 or iv_decimal <= 0 or dte_days <= 0:
        return 0.5
    t = dte_days / 365.0
    try:
        d1 = (math.log(underlying_price / strike) + 0.5 * iv_decimal**2 * t) / (iv_decimal * math.sqrt(t))
        d2 = d1 - iv_decimal * math.sqrt(t)
    except (ValueError, ZeroDivisionError):
        return 0.5
    if option_type.upper() == "PUT":
        # A put expires OTM if price stays above the strike.
        return _norm_cdf(d2)
    # A call expires OTM if price stays below the strike.
    return _norm_cdf(-d2)


def probability_itm(
    underlying_price: float, strike: float, iv_decimal: float, dte_days: int, option_type: str
) -> float:
    return 1.0 - probability_otm(underlying_price, strike, iv_decimal, dte_days, option_type)


def score_credit_spread(
    return_on_risk: float, prob_of_profit: float,
    avg_open_interest: float, max_open_interest: int, total_bid_ask_spread: float,
    credit_pct_of_width: float
) -> float:
    # The old "distance vs expected-move" term was a crude linear stand-in for
    # probability of profit. Replaced with an actual Black-Scholes POP (see
    # probability_otm), which uses the strike's own IV and DTE instead of a
    # blanket VIX-based distance heuristic.
    return_score = min(25.0, return_on_risk / 0.35 * 25.0)
    pop_score = min(30.0, max(0.0, prob_of_profit) / 0.85 * 30.0)
    oi_score = (avg_open_interest / max_open_interest * 20.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 15.0 - min(15.0, total_bid_ask_spread * 2.0))
    credit_score = min(10.0, credit_pct_of_width / 0.25 * 10.0)
    return round(return_score + pop_score + oi_score + tightness_score + credit_score, 2)


def build_opportunity(
    symbol: str, strategy: str, pcr: float, india_vix: float, underlying_price: float,
    buy_leg: dict[str, Any], sell_leg: dict[str, Any], max_open_interest: int, expiry: str,
    iv_decimal: float = 0.0,
) -> dict[str, Any] | None:
    buy_price = ask_price(buy_leg)
    sell_price = bid_price(sell_leg)
    net_debit = buy_price - sell_price
    spread_width = abs(float(sell_leg["strikePrice"]) - float(buy_leg["strikePrice"]))

    if net_debit <= 0 or spread_width <= net_debit:
        return None

    max_profit = spread_width - net_debit
    rr_ratio = max_profit / net_debit
    cost_efficiency = net_debit / spread_width
    avg_open_interest = (int(buy_leg.get("openInterest") or 0) + int(sell_leg.get("openInterest") or 0)) / 2
    bid_ask_spread = (ask_price(buy_leg) - bid_price(buy_leg) + ask_price(sell_leg) - bid_price(sell_leg))

    dte = days_to_expiry(expiry)
    iv_for_calc = iv_decimal if iv_decimal > 0 else max(0.01, india_vix / 100.0)
    if strategy == "Bull Call Spread":
        breakeven = float(buy_leg["strikePrice"]) + net_debit
        prob_of_profit = probability_itm(underlying_price, breakeven, iv_for_calc, dte, "CALL")
    else:  # Bear Put Spread
        breakeven = float(buy_leg["strikePrice"]) - net_debit
        prob_of_profit = probability_itm(underlying_price, breakeven, iv_for_calc, dte, "PUT")

    return {
        "Symbol": symbol,
        "Strategy": strategy,
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Buy Leg (Strike)": buy_leg["strikePrice"],
        "Sell Leg (Strike)": sell_leg["strikePrice"],
        "Net Debit": round(net_debit, 2),
        "Spread Width": round(spread_width, 2),
        "Max Profit": round(max_profit, 2),
        "Max Loss": round(net_debit, 2),
        "R:R Ratio": round(rr_ratio, 2),
        "Breakeven": round(breakeven, 2),
        "Probability of Profit": round(prob_of_profit, 3),
        "Avg OI": round(avg_open_interest),
        "Bid-Ask Spread": round(bid_ask_spread, 2),
        "Score": score_spread(
            rr_ratio, cost_efficiency, avg_open_interest, max_open_interest, bid_ask_spread, prob_of_profit
        ),
    }


def build_credit_spread_opportunity(
    symbol: str, strategy: str, pcr: float, india_vix: float, underlying_price: float,
    short_leg: dict[str, Any], long_leg: dict[str, Any], max_open_interest: int,
    expiry: str, config: ScannerConfig, iv_decimal: float = 0.0,
) -> dict[str, Any] | None:
    short_strike = float(short_leg["strikePrice"])
    long_strike = float(long_leg["strikePrice"])
    spread_width = abs(short_strike - long_strike)
    credit = bid_price(short_leg) - ask_price(long_leg)

    if credit <= 0 or spread_width <= credit:
        return None

    credit_pct_of_width = credit / spread_width
    if credit_pct_of_width < config.min_credit_pct_of_width:
        return None

    short_distance_pct = (underlying_price - short_strike) / underlying_price if "Put" in strategy else (short_strike - underlying_price) / underlying_price
    if short_distance_pct < config.min_short_distance_pct:
        return None

    max_loss = spread_width - credit
    return_on_risk = credit / max_loss
    avg_open_interest = (int(short_leg.get("openInterest") or 0) + int(long_leg.get("openInterest") or 0)) / 2
    bid_ask_spread = (ask_price(short_leg) - bid_price(short_leg) + ask_price(long_leg) - bid_price(long_leg))

    dte = days_to_expiry(expiry)
    iv_for_calc = iv_decimal if iv_decimal > 0 else max(0.01, india_vix / 100.0)
    option_type = "PUT" if "Put" in strategy else "CALL"
    prob_of_profit = probability_otm(underlying_price, short_strike, iv_for_calc, dte, option_type)

    return {
        "Symbol": symbol,
        "Strategy": strategy,
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": short_leg["strikePrice"],
        "Buy Leg (Strike)": long_leg["strikePrice"],
        "Credit": round(credit, 2),
        "Net Debit": 0,
        "Spread Width": round(spread_width, 2),
        "Max Profit": round(credit, 2),
        "Max Loss": round(max_loss, 2),
        "R:R Ratio": round(return_on_risk, 2),
        "Return on Risk": round(return_on_risk, 2),
        "Short Strike Distance %": round(short_distance_pct * 100, 2),
        "Expected Move %": round(expected_move_pct(iv_for_calc, expiry) * 100, 2),
        "Probability of Profit": round(prob_of_profit, 3),
        "Avg OI": round(avg_open_interest),
        "Bid-Ask Spread": round(bid_ask_spread, 2),
        "Score": score_credit_spread(
            return_on_risk, prob_of_profit,
            avg_open_interest, max_open_interest, bid_ask_spread, credit_pct_of_width
        ),
    }


def build_short_straddle_opportunity(
    symbol: str, pcr: float, india_vix: float, underlying_price: float,
    atm_call: dict[str, Any], atm_put: dict[str, Any],
    max_open_interest: int, expiry: str, iv_decimal: float,
) -> dict[str, Any] | None:
    """Short Straddle: sell ATM call + sell ATM put at same strike.
    Profits from time decay when the underlying stays near the strike.
    Best when IV is elevated (rich premium) and the market is rangebound."""
    call_strike = float(atm_call["strikePrice"])
    put_strike = float(atm_put["strikePrice"])
    if call_strike != put_strike:
        return None
    credit = bid_price(atm_call) + bid_price(atm_put)
    if credit <= 0:
        return None
    upper_be = call_strike + credit
    lower_be = put_strike - credit
    dte = days_to_expiry(expiry)
    # POP = probability price stays between both breakevens
    prob_above_lower = probability_itm(underlying_price, lower_be, iv_decimal, dte, "CALL")
    prob_below_upper = probability_otm(underlying_price, upper_be, iv_decimal, dte, "CALL")
    prob_of_profit = max(0.0, min(1.0, prob_above_lower + prob_below_upper - 1.0))

    avg_oi = (int(atm_call.get("openInterest") or 0) + int(atm_put.get("openInterest") or 0)) / 2
    bid_ask = (ask_price(atm_call) - bid_price(atm_call) + ask_price(atm_put) - bid_price(atm_put))
    # Unlimited risk — scale down score by premium collected vs current price
    premium_pct = credit / underlying_price
    pop_score = min(35.0, prob_of_profit / 0.6 * 35.0)
    premium_score = min(25.0, premium_pct * 500)
    oi_score = (avg_oi / max_open_interest * 25.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 15.0 - min(15.0, bid_ask * 2.0))
    score = round(pop_score + premium_score + oi_score + tightness_score, 2)

    return {
        "Symbol": symbol,
        "Strategy": "Short Straddle",
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": f"{call_strike} CE + {put_strike} PE",
        "Buy Leg (Strike)": None,
        "Credit": round(credit, 2),
        "Net Debit": 0,
        "Spread Width": None,
        "Max Profit": round(credit, 2),
        "Max Loss": "Unlimited",
        "R:R Ratio": None,
        "Return on Risk": None,
        "Breakeven": f"{round(upper_be, 2)} / {round(lower_be, 2)}",
        "Probability of Profit": round(prob_of_profit, 3),
        "Expected Move %": round(expected_move_pct(iv_decimal, expiry) * 100, 2),
        "Avg OI": round(avg_oi),
        "Bid-Ask Spread": round(bid_ask, 2),
        "Score": score,
    }


def build_long_strangle_opportunity(
    symbol: str, pcr: float, india_vix: float, underlying_price: float,
    otm_call: dict[str, Any], otm_put: dict[str, Any],
    max_open_interest: int, expiry: str, iv_decimal: float,
) -> dict[str, Any] | None:
    """Long Strangle: buy OTM call + buy OTM put at different strikes.
    Cheaper than a straddle but needs a larger move to profit.
    Best when IV is low and a large breakout/event is expected."""
    call_strike = float(otm_call["strikePrice"])
    put_strike = float(otm_put["strikePrice"])
    if call_strike <= underlying_price or put_strike >= underlying_price:
        return None
    if call_strike <= put_strike:
        return None
    call_cost = ask_price(otm_call)
    put_cost = ask_price(otm_put)
    total_debit = call_cost + put_cost
    if total_debit <= 0:
        return None
    upper_be = call_strike + total_debit
    lower_be = put_strike - total_debit
    dte = days_to_expiry(expiry)
    expected_move = expected_move_pct(iv_decimal, expiry) * underlying_price
    prob_up = probability_itm(underlying_price, upper_be, iv_decimal, dte, "CALL")
    prob_down = probability_itm(underlying_price, lower_be, iv_decimal, dte, "PUT")
    prob_of_profit = min(1.0, prob_up + prob_down)

    avg_oi = (int(otm_call.get("openInterest") or 0) + int(otm_put.get("openInterest") or 0)) / 2
    bid_ask = (ask_price(otm_call) - bid_price(otm_call) + ask_price(otm_put) - bid_price(otm_put))
    move_efficiency = min(1.0, expected_move / total_debit) if total_debit > 0 else 0.0
    pop_score = min(35.0, prob_of_profit / 0.5 * 35.0)
    move_score = min(30.0, move_efficiency * 30.0)
    oi_score = (avg_oi / max_open_interest * 20.0) if max_open_interest > 0 else 0.0
    tightness_score = max(0.0, 15.0 - min(15.0, bid_ask * 2.0))
    score = round(pop_score + move_score + oi_score + tightness_score, 2)

    return {
        "Symbol": symbol,
        "Strategy": "Long Strangle",
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": None,
        "Buy Leg (Strike)": f"{call_strike} CE + {put_strike} PE",
        "Net Debit": round(total_debit, 2),
        "Credit": 0,
        "Spread Width": round(call_strike - put_strike, 2),
        "Max Profit": "Unlimited",
        "Max Loss": round(total_debit, 2),
        "R:R Ratio": None,
        "Breakeven": f"{round(upper_be, 2)} / {round(lower_be, 2)}",
        "Probability of Profit": round(prob_of_profit, 3),
        "Expected Move %": round(expected_move_pct(iv_decimal, expiry) * 100, 2),
        "Avg OI": round(avg_oi),
        "Bid-Ask Spread": round(bid_ask, 2),
        "Score": score,
    }


def build_iron_condor_opportunity(
    symbol: str, pcr: float, india_vix: float, underlying_price: float,
    put_spread: dict[str, Any], call_spread: dict[str, Any], expiry: str
) -> dict[str, Any] | None:
    put_width = float(put_spread["Spread Width"])
    call_width = float(call_spread["Spread Width"])
    if put_width != call_width:
        return None

    credit = float(put_spread["Credit"]) + float(call_spread["Credit"])
    max_loss = put_width - credit
    if credit <= 0 or max_loss <= 0:
        return None

    avg_score = (float(put_spread["Score"]) + float(call_spread["Score"])) / 2
    return_on_risk = credit / max_loss
    distance_pct = min(float(put_spread["Short Strike Distance %"]), float(call_spread["Short Strike Distance %"]))
    # The condor only loses if EITHER side is breached, so joint POP is
    # approximately pop_put + pop_call - 1 (independence assumption), clamped to [0, 1].
    pop_put = float(put_spread.get("Probability of Profit", 0.5))
    pop_call = float(call_spread.get("Probability of Profit", 0.5))
    prob_of_profit = max(0.0, min(1.0, pop_put + pop_call - 1.0))

    return {
        "Symbol": symbol,
        "Strategy": "Iron Condor",
        "Expiry": expiry,
        "PCR": round(pcr, 2),
        "India VIX": round(india_vix, 2),
        "Underlying Price": round(underlying_price, 2),
        "Sell Leg (Strike)": f"{put_spread['Sell Leg (Strike)']} PE / {call_spread['Sell Leg (Strike)']} CE",
        "Buy Leg (Strike)": f"{put_spread['Buy Leg (Strike)']} PE / {call_spread['Buy Leg (Strike)']} CE",
        "Credit": round(credit, 2),
        "Net Debit": 0,
        "Spread Width": round(put_width, 2),
        "Max Profit": round(credit, 2),
        "Max Loss": round(max_loss, 2),
        "R:R Ratio": round(return_on_risk, 2),
        "Return on Risk": round(return_on_risk, 2),
        "Short Strike Distance %": round(distance_pct, 2),
        "Expected Move %": max(put_spread.get("Expected Move %", 0), call_spread.get("Expected Move %", 0)),
        "Probability of Profit": round(prob_of_profit, 3),
        "Avg OI": round((float(put_spread["Avg OI"]) + float(call_spread["Avg OI"])) / 2),
        "Bid-Ask Spread": round(float(put_spread["Bid-Ask Spread"]) + float(call_spread["Bid-Ask Spread"]), 2),
        "Score": round(min(100.0, avg_score + min(10.0, return_on_risk * 12.0)), 2),
    }


def near_underlying(option: dict[str, Any], underlying_price: float, window_pct: float) -> bool:
    strike = float(option.get("strikePrice") or 0)
    return underlying_price * (1 - window_pct) <= strike <= underlying_price * (1 + window_pct)


# ---------------------------------------------------------------------------
# Core Analysis Routines
# ---------------------------------------------------------------------------

def _collect_tradeable(
    records: list[dict[str, Any]], key: str, config: ScannerConfig,
    underlying_price: float, window_pct: float,
) -> list[dict[str, Any]]:
    return sorted(
        [
            row[key] for row in records
            if key in row
            and option_is_tradeable(row[key], config)
            and near_underlying(row[key], underlying_price, window_pct)
        ],
        key=lambda opt: float(opt["strikePrice"]),
    )


def _build_put_credit_spreads(
    symbol: str, context: MarketContext, config: ScannerConfig,
    india_vix: float, iv_decimal: float, puts: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results = []
    for short_idx, short_put in enumerate(puts):
        if float(short_put["strikePrice"]) >= context.underlying_price:
            continue
        for long_put in puts[max(0, short_idx - config.max_legs_apart):short_idx]:
            opp = build_credit_spread_opportunity(
                symbol, "Bull Put Credit Spread", context.pcr, india_vix,
                context.underlying_price, short_put, long_put,
                context.max_open_interest, context.expiry, config, iv_decimal,
            )
            if opp:
                opp = add_validation_fields(opp, context, config)
            if opp:
                results.append(opp)
    return results


def _build_call_credit_spreads(
    symbol: str, context: MarketContext, config: ScannerConfig,
    india_vix: float, iv_decimal: float, calls: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results = []
    for short_idx, short_call in enumerate(calls):
        if float(short_call["strikePrice"]) <= context.underlying_price:
            continue
        for long_call in calls[short_idx + 1:min(len(calls), short_idx + config.max_legs_apart + 1)]:
            opp = build_credit_spread_opportunity(
                symbol, "Bear Call Credit Spread", context.pcr, india_vix,
                context.underlying_price, short_call, long_call,
                context.max_open_interest, context.expiry, config, iv_decimal,
            )
            if opp:
                opp = add_validation_fields(opp, context, config)
            if opp:
                results.append(opp)
    return results


def _build_debit_spreads(
    symbol: str, strategy: str, context: MarketContext, config: ScannerConfig,
    india_vix: float, iv_decimal: float, options: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results = []
    for buy_idx, buy_leg in enumerate(options):
        for sell_leg in options[buy_idx + 1:min(len(options), buy_idx + config.max_legs_apart + 1)]:
            opp = build_opportunity(
                symbol, strategy, context.pcr, india_vix,
                context.underlying_price, buy_leg, sell_leg,
                context.max_open_interest, context.expiry, iv_decimal,
            )
            if opp:
                opp = add_validation_fields(opp, context, config)
            if opp:
                results.append(opp)
    return results


def _best_iron_condors(
    symbol: str, context: MarketContext, config: ScannerConfig,
    india_vix: float, put_spreads: list[dict[str, Any]], call_spreads: list[dict[str, Any]],
    top_n: int,
) -> list[dict[str, Any]]:
    condors = []
    for ps in put_spreads[:5]:
        for cs in call_spreads[:5]:
            if float(ps["Spread Width"]) != float(cs["Spread Width"]):
                continue
            condor = build_iron_condor_opportunity(
                symbol, context.pcr, india_vix, context.underlying_price,
                ps, cs, context.expiry,
            )
            if condor:
                condor = add_validation_fields(condor, context, config)
            if condor:
                condors.append(condor)
    condors.sort(key=lambda x: x["Score"], reverse=True)
    return condors[:max(1, top_n)]


def analyze_symbol(
    session: Any, symbol: str, india_vix: float, config: ScannerConfig,
    chain_type: str = "auto", expiry: str | None = None,
    trend_map: dict[str, str] | None = None, event_map: dict[str, str] | None = None,
    top_n: int = 1,
    strategy_mode: str = "both",
) -> list[dict[str, Any]]:
    """Single entry point that routes to the right strategy family based on
    TREND first, PCR second.

    Routing logic
    ─────────────
    BULLISH trend → Bull Call Spread (buy) + Bull Put Credit Spread (sell)
    BEARISH trend → Bear Put Spread (buy)  + Bear Call Credit Spread (sell)
    SIDEWAYS trend → Long Strangle (buy vol cheap) + Short Straddle (sell vol rich) + Iron Condor (sell bounded)
    NEUTRAL / UNKNOWN trend → all of the above evaluated; PCR used as tiebreaker

    Previously the two analysis functions (analyze_stock_spreads /
    analyze_option_selling_strategies) both routed by PCR alone. Trend was
    computed and stored but never actually influenced which strategies were
    built — it only appeared in the validation pass/fail flag afterwards.
    """
    data = fetch_option_chain(session, symbol, chain_type, expiry)
    if not data or "records" not in data:
        return []
    context = build_market_context(data, symbol, expiry, trend_map or {}, event_map or {}, config)
    if not context:
        return []

    sym = symbol.upper()
    iv = effective_iv(context, india_vix)
    trend = context.trend.lower()

    calls = _collect_tradeable(context.records, "CE", config, context.underlying_price, config.atm_window_pct)
    puts  = _collect_tradeable(context.records, "PE", config, context.underlying_price, config.atm_window_pct)

    candidates: list[dict[str, Any]] = []

    # ── BULLISH ──────────────────────────────────────────────────────────────
    if trend in ("bullish", "neutral", "unknown"):
        if strategy_mode in ("buying", "both"):
            # Bull Call Spread: buy lower ATM call, sell higher OTM call
            bull_calls = sorted(calls, key=lambda o: float(o["strikePrice"]))
            candidates.extend(_build_debit_spreads(sym, "Bull Call Spread", context, config, india_vix, iv, bull_calls))

        if strategy_mode in ("selling", "both"):
            # Bull Put Credit Spread: sell OTM put, buy further OTM put
            candidates.extend(_build_put_credit_spreads(sym, context, config, india_vix, iv, puts))

    # ── BEARISH ──────────────────────────────────────────────────────────────
    if trend in ("bearish", "neutral", "unknown"):
        if strategy_mode in ("buying", "both"):
            # Bear Put Spread: buy higher ATM put, sell lower OTM put
            bear_puts = sorted(puts, key=lambda o: float(o["strikePrice"]), reverse=True)
            candidates.extend(_build_debit_spreads(sym, "Bear Put Spread", context, config, india_vix, iv, bear_puts))

        if strategy_mode in ("selling", "both"):
            # Bear Call Credit Spread: sell OTM call, buy further OTM call
            candidates.extend(_build_call_credit_spreads(sym, context, config, india_vix, iv, calls))

    # ── SIDEWAYS ─────────────────────────────────────────────────────────────
    if trend in ("sideways", "neutral", "unknown"):
        # ATM options for straddle/strangle
        if calls and puts:
            atm_call = min(calls, key=lambda o: abs(float(o["strikePrice"]) - context.underlying_price))
            atm_put  = min(puts,  key=lambda o: abs(float(o["strikePrice"]) - context.underlying_price))

            if strategy_mode in ("buying", "both"):
                # Long Strangle: buy OTM call + OTM put
                otm_calls = [o for o in calls if float(o["strikePrice"]) > context.underlying_price]
                otm_puts  = [o for o in puts  if float(o["strikePrice"]) < context.underlying_price]
                strangle_pool = []
                for oc in otm_calls[:config.max_legs_apart]:
                    for op in otm_puts[-config.max_legs_apart:]:
                        opp = build_long_strangle_opportunity(
                            sym, context.pcr, india_vix, context.underlying_price,
                            oc, op, context.max_open_interest, context.expiry, iv,
                        )
                        if opp:
                            opp = add_validation_fields(opp, context, config)
                        if opp:
                            strangle_pool.append(opp)
                strangle_pool.sort(key=lambda x: x["Score"], reverse=True)
                candidates.extend(strangle_pool[:max(1, top_n)])

            if strategy_mode in ("selling", "both"):
                # Short Straddle: sell ATM call + ATM put
                opp = build_short_straddle_opportunity(
                    sym, context.pcr, india_vix, context.underlying_price,
                    atm_call, atm_put, context.max_open_interest, context.expiry, iv,
                )
                if opp:
                    opp = add_validation_fields(opp, context, config)
                if opp:
                    candidates.append(opp)

                # Iron Condor: combine best put credit + best call credit spread
                put_spreads  = _build_put_credit_spreads(sym, context, config, india_vix, iv, puts)
                call_spreads = _build_call_credit_spreads(sym, context, config, india_vix, iv, calls)
                put_spreads.sort(key=lambda x: x["Score"], reverse=True)
                call_spreads.sort(key=lambda x: x["Score"], reverse=True)
                if put_spreads and call_spreads:
                    candidates.extend(_best_iron_condors(sym, context, config, india_vix, put_spreads, call_spreads, top_n))

    candidates.sort(key=lambda x: x["Score"], reverse=True)
    return candidates[:max(1, top_n)]


# Keep old names as thin wrappers so any external code referencing them still works
def analyze_stock_spreads(session, symbol, india_vix, config, chain_type="auto",
                          expiry=None, trend_map=None, event_map=None, top_n=1):
    return analyze_symbol(session, symbol, india_vix, config, chain_type, expiry,
                          trend_map, event_map, top_n, strategy_mode="buying")


def analyze_option_selling_strategies(session, symbol, india_vix, config, chain_type="auto",
                                      expiry=None, trend_map=None, event_map=None, top_n=1):
    return analyze_symbol(session, symbol, india_vix, config, chain_type, expiry,
                          trend_map, event_map, top_n, strategy_mode="selling")


def write_results(results: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(results).sort_values(["Score", "R:R Ratio"], ascending=False)

    if output_path.suffix.lower() == ".csv":
        df.to_csv(output_path, index=False)
        return

    # xlsx: write an "All Opportunities" sheet (everything, with failed-validation rows
    # highlighted red so they're visible at a glance instead of needing to be filtered
    # manually) plus a "Strict (enforce-validations)" sheet containing only the rows that
    # would survive --enforce-validations, pre-sorted by Score. This used to require a
    # separate manual post-processing pass on the output file; now every run gets both
    # views for free, regardless of whether --enforce-validations was passed on the CLI.
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    headers = list(df.columns)
    has_validation_col = "Validation Pass" in headers

    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Opportunities"

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    bold = Font(bold=True)

    def write_sheet(ws, frame: pd.DataFrame, highlight_failed: bool) -> None:
        for row in dataframe_to_rows(frame, index=False, header=True):
            ws.append(row)
        for cell in ws[1]:
            cell.font = bold
        if highlight_failed and has_validation_col:
            val_col_idx = headers.index("Validation Pass") + 1
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=val_col_idx).value is False:
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = red_fill
                        cell.font = red_font
        ws.freeze_panes = "A2"
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            values = [str(header)] + [str(v) for v in frame[header].tolist()]
            max_len = max(len(v) for v in values)
            ws.column_dimensions[col_letter].width = min(28, max(10, max_len + 2))

    write_sheet(ws_all, df, highlight_failed=True)

    if has_validation_col:
        strict_df = df[df["Validation Pass"] == True].sort_values(  # noqa: E712
            ["Score", "R:R Ratio"], ascending=False
        )
        ws_strict = wb.create_sheet("Strict (enforce-validations)")
        write_sheet(ws_strict, strict_df, highlight_failed=False)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Integrated Scanner: Technical Trend + NSE Option Chain Spread Analysis.")
    parser.add_argument("--symbols", nargs="+", default=list(DEFAULT_SYMBOLS))
    parser.add_argument("--output", default="Combined_Option_Spread_Analysis.xlsx")
    parser.add_argument("--min-oi", type=int, default=100)
    parser.add_argument("--min-bid", type=float, default=0.05)
    parser.add_argument("--atm-window-pct", type=float, default=0.12)
    parser.add_argument("--max-legs-apart", type=int, default=8)
    parser.add_argument("--expiry", help="NSE expiry format e.g. 25-Aug-2026")
    parser.add_argument("--chain-type", choices=("auto", "equity", "indices"), default="auto")
    parser.add_argument("--strategy-mode", choices=("buying", "selling", "both"), default="both")
    parser.add_argument("--min-credit-pct-of-width", type=float, default=0.12)
    parser.add_argument("--min-short-distance-pct", type=float, default=0.01)
    parser.add_argument("--max-bid-ask-spread-pct", type=float, default=0.25)
    parser.add_argument("--min-volume", type=int, default=10)
    parser.add_argument("--top-n", type=int, default=1, help="Number of ranked candidates to keep per symbol/strategy-type instead of just the single best.")
    parser.add_argument("--lot-size", type=int, default=1)
    parser.add_argument("--max-margin-per-trade", type=float)
    parser.add_argument("--trend-file", help="CSV with Symbol,Trend columns.")
    parser.add_argument("--event-file", help="CSV with Symbol,EventRisk columns.")
    parser.add_argument("--enforce-validations", action="store_true")
    parser.add_argument(
        "--trend-debug", action="store_true",
        help="Also write a <output>_trend_debug.csv with raw indicator values and each "
             "bullish/bearish/sideways condition's pass/fail per symbol, so a 'neutral' "
             "result can be inspected instead of trusted blindly."
    )
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--cookie-header")
    parser.add_argument("--cookie-file")
    parser.add_argument("--browser-impersonation", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO if args.verbose else logging.WARNING, format="%(levelname)s: %(message)s")

    config = ScannerConfig(
        min_bid=args.min_bid,
        min_open_interest=args.min_oi,
        max_legs_apart=args.max_legs_apart,
        atm_window_pct=args.atm_window_pct,
        min_credit_pct_of_width=args.min_credit_pct_of_width,
        min_short_distance_pct=args.min_short_distance_pct,
        max_bid_ask_spread_pct=args.max_bid_ask_spread_pct,
        min_volume=args.min_volume,
        max_margin_per_trade=args.max_margin_per_trade,
        lot_size=args.lot_size,
        enforce_validations=args.enforce_validations,
    )

    trend_map = load_symbol_map(args.trend_file, "Trend")
    event_map = load_symbol_map(args.event_file, "EventRisk")
    cookie_header = load_cookie_header(args.cookie_header, args.cookie_file)

    try:
        session = create_nse_session(cookie_header, args.browser_impersonation)
    except Exception as exc:
        raise SystemExit(f"Unable to initialize NSE session: {exc}") from exc

    india_vix = fetch_india_vix(session, config.fallback_vix)
    print(f"Current India VIX: {india_vix:.2f}")

    trend_diagnostics: list[dict[str, Any]] = []

    results: list[dict[str, Any]] = []
    for symbol in args.symbols:
        print(f"Analyzing {symbol.upper()}...")

        if args.trend_debug:
            _, diag = detect_market_trend_detailed(symbol.upper(), lookback_period=config.lookback_period)
            trend_diagnostics.append(diag)

        results.extend(
            analyze_symbol(
                session, symbol, india_vix, config,
                chain_type=args.chain_type, expiry=args.expiry,
                trend_map=trend_map, event_map=event_map,
                top_n=args.top_n, strategy_mode=args.strategy_mode,
            )
        )

    if args.trend_debug and trend_diagnostics:
        debug_path = Path(args.output).with_name(Path(args.output).stem + "_trend_debug.csv")
        pd.DataFrame(trend_diagnostics).to_csv(debug_path, index=False)
        print(f"Trend diagnostics saved to '{debug_path}'.")

    if not results:
        print("No qualifying spread opportunities found.")
        return

    output_path = Path(args.output)
    write_results(results, output_path)
    print(f"Analysis completed. Output saved to '{output_path}'.")


if __name__ == "__main__":
    main()